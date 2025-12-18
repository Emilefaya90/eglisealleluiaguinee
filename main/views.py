from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET, require_http_methods
from django.views.decorators.csrf import ensure_csrf_cookie
from django.conf import settings
from .models import CulteEffectif
import json
from datetime import datetime
from .models import Eglise, PersonnelPastoral, Membre, Departement, DepartementAttachment, Finance, FinanceReport, ComptabiliteReport, FinanceAttachment, StrategicPlan
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.core.mail import send_mail
import re
from .forms import MembreForm
from datetime import datetime
from decimal import Decimal, InvalidOperation
import os
import json
from django.core.serializers.json import DjangoJSONEncoder
import unicodedata as _ud
from django.forms.models import model_to_dict
from io import BytesIO

# Helpers souples pour sérialisation
def _get_first_attr(obj, names, default=None):
    for n in names:
        if hasattr(obj, n):
            v = getattr(obj, n)
            try:
                # FileField a .url
                if hasattr(v, 'url'):
                    return v
            except Exception:
                pass
            if v is not None:
                return v
    return default

def _to_str(v):
    try:
        return str(v)
    except Exception:
        return ''

def _file_url_from_any(v):
    # v peut être une FileField (avec .url) ou une chaîne chemin relatif
    try:
        if hasattr(v, 'url'):
            return v.url
    except Exception:
        pass
    s = _to_str(v)
    if not s:
        return ''
    if s.startswith('http://') or s.startswith('https://') or s.startswith('/'):
        return s
    media = getattr(settings, 'MEDIA_URL', '/media/')
    if not media.endswith('/'):
        media += '/'
    return media + s.lstrip('/')

# Tentative de correction de mojibake (UTF-8 interprété en Latin-1)
def _fix_mojibake(text: str) -> str:
    if not isinstance(text, str):
        try:
            text = str(text)
        except Exception:
            return ''
    # Si on voit des séquences typiques de mojibake
    if 'Ã' in text or 'Â' in text or '¤' in text:
        try:
            # essayer latin-1 -> bytes -> utf-8
            repaired = text.encode('latin-1', errors='ignore').decode('utf-8', errors='ignore')
            if repaired:
                return repaired
        except Exception:
            pass
    return text

def _serialize_attachments_for_report(rapport):
    items = []
    # 1) Relation directe si présente: report.attachments
    att_rel = getattr(rapport, 'attachments', None)
    att_iter = None
    if att_rel is not None:
        try:
            att_iter = att_rel.all()
        except Exception:
            att_iter = None
    # 2) Sinon, modèle FinanceAttachment si importé
    if att_iter is None:
        try:
            att_iter = FinanceAttachment.objects.filter(report=rapport)
        except Exception:
            att_iter = []
    for a in att_iter or []:
        file_any = _get_first_attr(a, ['file', 'fichier', 'document', 'path', 'url'], '')
        url = _file_url_from_any(file_any)
        name = _get_first_attr(a, ['name', 'filename', 'titre', 'title'], '')
        if not name:
            try:
                import os
                name = os.path.basename(url) or 'PJ'
            except Exception:
                name = 'PJ'
        if url:
            items.append({'url': url, 'name': _to_str(name)})
    return items

def _aggregate_finance_lines_for_report(rapport):
    """Retourne un dict avec CSV agrégés pour revenus et dépenses en essayant plusieurs noms de champs."""
    try:
        lignes = Finance.objects.filter(report_id=rapport.id).order_by('date_operation', 'id')
    except Exception:
        try:
            lignes = Finance.objects.filter(report=rapport).order_by('id')
        except Exception:
            lignes = []
    rev_sources, rev_montants, rev_dates, rev_modes = [], [], [], []
    dep_natures, dep_montants, dep_benef, dep_justifs, dep_cats = [], [], [], [], []
    for li in lignes:
        # Champs possibles
        montant = _get_first_attr(li, ['montant', 'amount', 'valeur'], 0) or 0
        try:
            from decimal import Decimal
            montant_num = Decimal(str(montant))
        except Exception:
            try:
                montant_num = Decimal(str(float(montant)))
            except Exception:
                montant_num = Decimal('0')
        dateop = _get_first_attr(li, ['date_operation', 'date', 'created_at', 'date_op'], '') or ''
        mode = _get_first_attr(li, ['mode_paiement', 'mode', 'payment_method'], '') or ''
        source = _get_first_attr(li, ['source', 'source_revenu', 'libelle', 'intitule'], '') or ''
        nature = _get_first_attr(li, ['nature', 'nature_depense', 'objet'], '') or ''
        beneficiaire = _get_first_attr(li, ['beneficiaire', 'destinataire', 'tiers'], '') or ''
        justificatif = _get_first_attr(li, ['justificatif', 'piece', 'reference'], '') or ''
        categorie = _get_first_attr(li, ['categorie', 'category'], '') or ''
        typeop = _to_str(_get_first_attr(li, ['type_operation', 'type', 'sens'], ''))
        typeop_l = typeop.lower()
        is_rev = ('recette' in typeop_l) or ('revenu' in typeop_l) or (('depense' not in typeop_l) and (montant_num >= 0))
        is_dep = ('depense' in typeop_l) or (montant_num < 0)
        if is_rev:
            rev_sources.append(_to_str(source))
            rev_montants.append(str(abs(montant_num)))
            rev_dates.append(_to_str(getattr(dateop, 'strftime', lambda fmt: dateop)('%Y-%m-%d') if hasattr(dateop, 'strftime') else dateop))
            rev_modes.append(_to_str(mode))
        elif is_dep:
            dep_natures.append(_to_str(nature))
            dep_montants.append(str(abs(montant_num)))
            dep_benef.append(_to_str(beneficiaire))
            dep_justifs.append(_to_str(justificatif))
            dep_cats.append(_to_str(categorie))
    return {
        'sources_revenus': ', '.join([x for x in rev_sources if x]),
        'montants_revenus': ', '.join([x for x in rev_montants if x]),
        'dates_revenus': ', '.join([x for x in rev_dates if x]),
        'modes_paiement_revenus': ', '.join([x for x in rev_modes if x]),
        'natures_depenses': ', '.join([x for x in dep_natures if x]),
        'montants_depenses': ', '.join([x for x in dep_montants if x]),
        'beneficiaires': ', '.join([x for x in dep_benef if x]),
        'justificatifs': ', '.join([x for x in dep_justifs if x]),
        'categories': ', '.join([x for x in dep_cats if x]),
    }

def accueil(request):
    """Vue pour la page d'accueil"""
    return render(request, 'main/accueil.html')

def apropos(request):
    """Vue pour la page À propos"""
    return render(request, 'main/apropos.html')

def historique(request):
    return render(request, 'main/historique.html')

def objectif_general(request):
    return render(request, 'main/objectif_general.html')

def objectif_specifique(request):
    return render(request, 'main/objectif_specifique.html')

@login_required
def gestion(request):
    """Vue pour la page Gestion"""
    return render(request, 'main/gestion.html')

@login_required
def eglises(request):
    """Vue pour la page Eglises"""
    # Récupérer toutes les églises pour le tableau
    eglises = Eglise.objects.all().order_by('-date_creation')
    
    # Dictionnaire pour stocker les données du formulaire en cas d'erreur
    form_data = {}
    # PRG: restaurer d'éventuelles données de formulaire depuis la session
    try:
        stored = request.session.pop('eglise_form_data', None)
        if stored:
            form_data = stored
    except Exception:
        pass
    
    if request.method == 'POST':
        # Gestion de la suppression d'église
        if request.POST.get('action') == 'supprimer':
            try:
                eglise_id = request.POST.get('eglise_id')
                eglise = Eglise.objects.get(id=eglise_id)
                nom_eglise = eglise.nom
                eglise.delete()
                messages.success(request, f"L'église '{nom_eglise}' a été supprimée avec succès.")
                return redirect('main:eglises')
            except Eglise.DoesNotExist:
                messages.error(request, "L'église demandée n'existe pas.")
                return redirect('main:eglises')
            except Exception as e:
                # Traduire l'erreur en français
                error_msg = str(e)
                if 'got unexpected keyword arguments' in error_msg:
                    error_msg = 'Arguments inattendus dans le formulaire. Veuillez vérifier les champs.'
                return render(request, 'main/eglises.html', {'error': error_msg, 'form_data': form_data})
        else:
            # Stocker toutes les données du formulaire (pour ré-affichage en cas d'erreur)
            form_data = {
                'nom': request.POST.get('nom', ''),
                'adresse': request.POST.get('adresse', ''),
                'ville': request.POST.get('ville', ''),
                'quartier': request.POST.get('quartier', ''),
                'pays': request.POST.get('pays', 'Guinée'),
                'date_creation': request.POST.get('date_creation', ''),
                'email': request.POST.get('email', ''),
                'telephone': request.POST.get('telephone', ''),
                'responsable': request.POST.get('responsable', ''),
                'nombre_membres': request.POST.get('nombre_membres', '0'),
                'est_association': request.POST.get('est_association', ''),
                'numero_autorisation': request.POST.get('numero_autorisation', ''),
                'date_enregistrement': request.POST.get('date_enregistrement', ''),
                'autres_activites_detail': request.POST.get('autres_activites_detail', ''),
                'nom_declarant': request.POST.get('nom_declarant', ''),
                'qualite_declarant': request.POST.get('qualite_declarant', ''),
                'date_declaration': request.POST.get('date_declaration', ''),
                'activites': request.POST.getlist('activites[]', [])
            }
            
            # Validation des champs obligatoires
            required_fields = ['nom', 'ville', 'pays', 'date_creation', 'responsable']
            missing_fields = []
            for field in required_fields:
                if not request.POST.get(field):
                    missing_fields.append(field)
            if missing_fields:
                messages.error(request, "Remplissez correctement les champs obligatoires.")
                # PRG: conserver form_data et rediriger
                try:
                    request.session['eglise_form_data'] = form_data
                except Exception:
                    pass
                return redirect('main:eglises')
            
            try:
                # Helpers
                def to_bool(val: str) -> bool:
                    v = (val or '').strip().lower()
                    return v in ('oui', 'true', '1', 'on')

                # Convertir la liste des activités en CSV (TextField)
                activites_csv = ''
                try:
                    activites_list = form_data.get('activites') or []
                    if isinstance(activites_list, list):
                        activites_csv = ','.join([str(x) for x in activites_list if str(x).strip()])
                    else:
                        activites_csv = str(activites_list)
                except Exception:
                    activites_csv = ''

                # Créer une nouvelle église
                nouvelle_eglise = Eglise(
                    nom=form_data['nom'],
                    adresse=form_data['adresse'],
                    ville=form_data['ville'],
                    quartier=form_data['quartier'],
                    pays=form_data['pays'],
                    date_creation=datetime.strptime(form_data['date_creation'], '%Y-%m-%d').date(),
                    email=form_data['email'],
                    telephone=form_data['telephone'],
                    responsable=form_data['responsable'],
                    nombre_membres=int(form_data['nombre_membres']) if form_data['nombre_membres'] else 0,
                    est_association=to_bool(form_data['est_association']),
                    numero_autorisation=form_data['numero_autorisation'],
                    date_enregistrement=datetime.strptime(form_data['date_enregistrement'], '%Y-%m-%d').date() if form_data['date_enregistrement'] else None,
                    activites=activites_csv,
                    autres_activites_detail=form_data['autres_activites_detail'],
                    nom_declarant=form_data['nom_declarant'],
                    qualite_declarant=form_data['qualite_declarant'],
                    date_declaration=datetime.strptime(form_data['date_declaration'], '%Y-%m-%d').date() if form_data['date_declaration'] else None
                )

                # Gérer les pièces jointes si fournies
                if 'statuts_doc' in request.FILES:
                    nouvelle_eglise.statuts_doc = request.FILES['statuts_doc']
                if 'pieces_identite_doc' in request.FILES:
                    nouvelle_eglise.pieces_identite_doc = request.FILES['pieces_identite_doc']

                nouvelle_eglise.save()

                messages.success(request, f"L'église '{form_data['nom']}' a été enregistrée avec succès.")
                return redirect('main:eglises')
                
            except ValueError as e:
                if "time data" in str(e):
                    messages.error(request, "Format de date invalide. Utilisez le format AAAA-MM-JJ.")
                else:
                    messages.error(request, f"Erreur de validation : {str(e)}")
                # PRG sur erreur: stocker et rediriger
                try:
                    request.session['eglise_form_data'] = form_data
                except Exception:
                    pass
                return redirect('main:eglises')
            except Exception as e:
                messages.error(request, f"Erreur lors de l'enregistrement : {str(e)}")
                try:
                    request.session['eglise_form_data'] = form_data
                except Exception:
                    pass
                return redirect('main:eglises')
    
    return render(request, 'main/eglises.html', {
        'eglises': eglises,
        'form_data': form_data
    })

def ajout_eglise(request):
    """Vue pour ajouter une nouvelle église"""
    return render(request, 'main/ajout_eglise.html')

def modifier_eglise(request, eglise_id):
    """Vue pour modifier une église existante"""
    eglise = get_object_or_404(Eglise, id=eglise_id)
    
    if request.method == 'POST':
        # Traitement de la modification
        pass
    
    return render(request, 'main/modifier_eglise.html', {'eglise': eglise})

def visite_pastorale(request):
    """Vue pour la gestion des visites pastorales"""
    return render(request, 'main/visite_pastorale.html')

def projet_local(request):
    """Vue pour la gestion des projets locaux"""
    return render(request, 'main/projet_local.html')

from django.contrib import messages

@login_required
def personnel_pastoral(request):
    """Vue pour la gestion du personnel pastoral"""
    # Récupérer tous les pasteurs et églises pour affichage
    pasteurs = PersonnelPastoral.objects.all().order_by('-id')
    eglises = Eglise.objects.all().order_by('nom')
    
    if request.method == 'POST':
        try:
            # Récupération des données du formulaire
            # Section 1: Informations personnelles
            nom = request.POST.get('nom', '').strip()
            prenom = request.POST.get('prenoms', '').strip()
            fonction = request.POST.get('fonction', '').strip()
            sexe = request.POST.get('sexe', '').strip()
            date_naissance = request.POST.get('date_naissance')
            lieu_naissance = request.POST.get('lieu_naissance', '').strip()
            nationalite = request.POST.get('nationalite', '').strip()
            domicile = request.POST.get('domicile', '').strip()
            etat_civil = request.POST.get('etat_civil', '').strip()
            nombre_enfants = request.POST.get('nombre_enfants', '0')
            profession = request.POST.get('profession', '').strip()
            telephone = request.POST.get('telephone', '').strip()
            email = request.POST.get('email', '').strip()
            date_consecration = request.POST.get('date_consecration')
            lieu_consecration = request.POST.get('lieu_consecration', '').strip()
            consacre_par = request.POST.get('consacre_par', '').strip()
            
            # Section 2: Filiations
            prenoms_pere = request.POST.get('prenoms_pere', '').strip()
            prenoms_nom_mere = request.POST.get('prenoms_nom_mere', '').strip()
            
            # Section 3: Affectation
            eglise_id = request.POST.get('eglise_affectee')
            lieu_affectation = request.POST.get('lieu_affectation', '').strip()
            date_affectation = request.POST.get('date_affectation')
            region = request.POST.get('region', '').strip()
            zone = request.POST.get('zone', '').strip()
            
            # Section 4: Formations
            types_formations = request.POST.get('types_formations', '').strip()
            statut_actuel = request.POST.get('statut_actuel', '').strip()
            
            # Validation des champs obligatoires
            champs_obligatoires = {
                'nom': 'Nom',
                'prenom': 'Prénom',
                'fonction': 'Fonction',
                'sexe': 'Sexe',
                'eglise_id': 'Église d\'affectation'
            }
            
            champs_manquants = []
            champs_values = {
                'nom': nom,
                'prenom': prenom,
                'fonction': fonction,
                'sexe': sexe,
                'eglise_id': eglise_id
            }
            
            for champ, nom_champ in champs_obligatoires.items():
                if not champs_values[champ]:
                    champs_manquants.append(nom_champ)
            
            if champs_manquants:
                if len(champs_manquants) == 1:
                    messages.error(request, f"Veuillez remplir le champ obligatoire : {champs_manquants[0]}")
                else:
                    messages.error(request, f"Veuillez remplir les champs obligatoires suivants : {', '.join(champs_manquants)}")
                
                # Conserver les données saisies pour éviter de vider le formulaire
                form_data = {
                    'nom': nom,
                    'prenom': prenom,
                    'fonction': fonction,
                    'sexe': sexe,
                    'date_naissance': date_naissance,
                    'lieu_naissance': lieu_naissance,
                    'nationalite': nationalite,
                    'domicile': domicile,
                    'etat_civil': etat_civil,
                    'nombre_enfants': nombre_enfants,
                    'profession': profession,
                    'telephone': telephone,
                    'email': email,
                    'date_consecration': date_consecration,
                    'lieu_consecration': lieu_consecration,
                    'consacre_par': consacre_par,
                    'prenoms_pere': prenoms_pere,
                    'prenoms_nom_mere': prenoms_nom_mere,
                    'eglise_affectee': eglise_id,
                    'lieu_affectation': lieu_affectation,
                    'date_affectation': date_affectation,
                    'region': region,
                    'zone': zone,
                    'types_formations': types_formations,
                    'statut_actuel': statut_actuel,
                }
                
                return render(request, 'main/personnel_pastoral.html', {
                    'pasteurs': pasteurs,
                    'eglises': eglises,
                    'form_data': form_data,
                })
            
            # Vérifier que l'église existe
            try:
                eglise = Eglise.objects.get(id=int(eglise_id))
            except (ValueError, Eglise.DoesNotExist):
                messages.error(request, "L'église sélectionnée n'existe pas. Veuillez choisir une église valide dans la liste.")
                
                # Conserver les données saisies
                form_data = {
                    'nom': nom,
                    'prenom': prenom,
                    'fonction': fonction,
                    'sexe': sexe,
                    'date_naissance': date_naissance,
                    'lieu_naissance': lieu_naissance,
                    'nationalite': nationalite,
                    'domicile': domicile,
                    'etat_civil': etat_civil,
                    'nombre_enfants': nombre_enfants,
                    'profession': profession,
                    'telephone': telephone,
                    'email': email,
                    'date_consecration': date_consecration,
                    'lieu_consecration': lieu_consecration,
                    'consacre_par': consacre_par,
                    'prenoms_pere': prenoms_pere,
                    'prenoms_nom_mere': prenoms_nom_mere,
                    'eglise_affectee': eglise_id,
                    'lieu_affectation': lieu_affectation,
                    'date_affectation': date_affectation,
                    'region': region,
                    'zone': zone,
                    'types_formations': types_formations,
                    'statut_actuel': statut_actuel,
                }
                
                return render(request, 'main/personnel_pastoral.html', {
                    'pasteurs': pasteurs,
                    'eglises': eglises,
                    'form_data': form_data,
                })
            
            # Créer et enregistrer le personnel pastoral
            PersonnelPastoral.objects.create(
                nom=nom,

                fonction=fonction,
                sexe=sexe,
                date_naissance=datetime.strptime(date_naissance, '%Y-%m-%d').date() if date_naissance else None,
                lieu_naissance=lieu_naissance,
                nationalite=nationalite,
                domicile=domicile,
                etat_civil=etat_civil,
                nombre_enfants=int(nombre_enfants) if nombre_enfants else 0,
                profession=profession,
                telephone=telephone,
                email=email,
                date_consecration=datetime.strptime(date_consecration, '%Y-%m-%d').date() if date_consecration else None,
                lieu_consecration=lieu_consecration,
                consacre_par=consacre_par,
                prenom=prenom,
                prenoms_pere=prenoms_pere,
                prenoms_nom_mere=prenoms_nom_mere,
                eglise=eglise,
                lieu_affectation=lieu_affectation,
                date_affectation=datetime.strptime(date_affectation, '%Y-%m-%d').date() if date_affectation else None,
                region=region,
                zone=zone,
                types_formations=types_formations,
                statut_actuel=statut_actuel,
                photo=request.FILES.get('photo'),
                document_fichier=request.FILES.get('document_fichier')
            )
            
            # Redirection POST-Redirect-GET avec message de succès
            messages.success(request, "Enregistrement effectué avec succès.")
            return redirect('/personnel-pastoral/')
            
        except ValueError as e:
            if "invalid literal for int()" in str(e):
                messages.error(request, "Erreur dans les données numériques. Vérifiez le nombre d'enfants et l'église sélectionnée.")
            elif "time data" in str(e):
                messages.error(request, "Format de date invalide. Utilisez le format JJ/MM/AAAA ou AAAA-MM-JJ.")
            else:
                messages.error(request, f"Erreur de validation des données : {str(e)}")
        except Eglise.DoesNotExist:
            messages.error(request, "L'église sélectionnée n'existe pas dans le système. Veuillez choisir une église valide.")
        except Exception as e:
            error_message = str(e)
            # Traduire les messages d'erreur courants en français
            if "Field 'id' expected a number but got" in error_message:
                messages.error(request, "Erreur : l'église sélectionnée n'est pas valide. Veuillez choisir une église dans la liste.")
            elif "UNIQUE constraint failed" in error_message:
                messages.error(request, "Ce pasteur existe déjà dans le système. Veuillez vérifier les informations saisies.")
            else:
                # Supprimer le message verbeux de reverse pour satisfaire la demande
                if "Reverse for" in error_message:
                    pass  # Ne rien afficher
                else:
                    messages.error(request, f"Une erreur s'est produite lors de l'enregistrement.")
    
    return render(request, 'main/personnel_pastoral.html', {
        'pasteurs': pasteurs,
        'eglises': eglises
    })

def get_pasteur_details(request, pasteur_id):
    """Vue pour récupérer les détails d'un pasteur en format JSON"""
    try:
        pasteur = get_object_or_404(PersonnelPastoral, id=pasteur_id)
        
        data = {
            'id': pasteur.id,
            'nom': pasteur.nom,
            'prenom': pasteur.prenom,
            'fonction': pasteur.fonction,
            'sexe': pasteur.sexe,
            'date_naissance': pasteur.date_naissance.strftime('%Y-%m-%d') if pasteur.date_naissance else '',
            'lieu_naissance': pasteur.lieu_naissance,
            'nationalite': pasteur.nationalite,
            'domicile': pasteur.domicile,
            'etat_civil': pasteur.etat_civil,
            'nombre_enfants': pasteur.nombre_enfants,
            'profession': pasteur.profession,
            'telephone': pasteur.telephone,
            'email': pasteur.email,
            'date_consecration': pasteur.date_consecration.strftime('%Y-%m-%d') if pasteur.date_consecration else '',
            'lieu_consecration': pasteur.lieu_consecration,
            'consacre_par': pasteur.consacre_par,
            'prenoms_pere': pasteur.prenoms_pere,
            'prenoms_nom_mere': pasteur.prenoms_nom_mere,
            'eglise_id': pasteur.eglise.id if pasteur.eglise else '',
            'eglise_nom': pasteur.eglise.nom if pasteur.eglise else '',
            'lieu_affectation': pasteur.lieu_affectation,
            'date_affectation': pasteur.date_affectation.strftime('%Y-%m-%d') if pasteur.date_affectation else '',
            'region': pasteur.region,
            'zone': pasteur.zone,
            'types_formations': pasteur.types_formations,
            'statut_actuel': pasteur.statut_actuel,
            'photo_url': pasteur.photo.url if pasteur.photo else '',
        }
        
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

# --- Actions sur un pasteur (modifier / supprimer) ---
def modifier_pasteur(request, pasteur_id: int):
    """Met à jour un pasteur. Retour JSON pour les requêtes AJAX, sinon redirection.

    Attend les champs du formulaire d'édition (voir modal `modal_edit_pasteur.html`).
    """
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    try:
        pasteur = get_object_or_404(PersonnelPastoral, id=pasteur_id)
        if request.method != 'POST':
            if is_ajax:
                return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'}, status=405)
            messages.error(request, "Méthode non autorisée pour la modification.")
            return redirect('/personnel-pastoral/')
        # Mise à jour des champs simples
        pasteur.prenom = request.POST.get('prenom', pasteur.prenom)
        pasteur.nom = request.POST.get('nom', pasteur.nom)
        pasteur.fonction = request.POST.get('fonction', pasteur.fonction)
        pasteur.sexe = request.POST.get('sexe', pasteur.sexe)
        pasteur.lieu_naissance = request.POST.get('lieu_naissance', pasteur.lieu_naissance)
        pasteur.nationalite = request.POST.get('nationalite', pasteur.nationalite)
        pasteur.domicile = request.POST.get('domicile', pasteur.domicile)
        pasteur.etat_civil = request.POST.get('etat_civil', pasteur.etat_civil)
        pasteur.nombre_enfants = request.POST.get('nombre_enfants', pasteur.nombre_enfants)
        pasteur.profession = request.POST.get('profession', pasteur.profession)
        pasteur.telephone = request.POST.get('telephone', pasteur.telephone)
        pasteur.email = request.POST.get('email', pasteur.email)
        pasteur.lieu_consecration = request.POST.get('lieu_consecration', pasteur.lieu_consecration)
        pasteur.consacre_par = request.POST.get('consacre_par', pasteur.consacre_par)
        pasteur.prenoms_pere = request.POST.get('prenoms_pere', pasteur.prenoms_pere)
        pasteur.prenoms_nom_mere = request.POST.get('prenoms_nom_mere', pasteur.prenoms_nom_mere)
        pasteur.lieu_affectation = request.POST.get('lieu_affectation', pasteur.lieu_affectation)
        pasteur.region = request.POST.get('region', pasteur.region)
        pasteur.zone = request.POST.get('zone', pasteur.zone)
        pasteur.types_formations = request.POST.get('types_formations', pasteur.types_formations)
        pasteur.statut_actuel = request.POST.get('statut_actuel', pasteur.statut_actuel)
        pasteur.diplomes_obtenus = request.POST.get('diplomes_obtenus', getattr(pasteur, 'diplomes_obtenus', ''))

        # Dates (format YYYY-MM-DD)
        date_naissance = request.POST.get('date_naissance')
        if date_naissance:
            try:
                pasteur.date_naissance = datetime.strptime(date_naissance, '%Y-%m-%d').date()
            except Exception:
                pass
        date_consecration = request.POST.get('date_consecration')
        if date_consecration:
            try:
                pasteur.date_consecration = datetime.strptime(date_consecration, '%Y-%m-%d').date()
            except Exception:
                pass
        date_affectation = request.POST.get('date_affectation')
        if date_affectation:
            try:
                pasteur.date_affectation = datetime.strptime(date_affectation, '%Y-%m-%d').date()
            except Exception:
                pass

        # Église (FK)
        eglise_id = request.POST.get('eglise') or request.POST.get('eglise_id')
        if eglise_id:
            try:
                pasteur.eglise = Eglise.objects.get(id=eglise_id)
            except Exception:
                pasteur.eglise = pasteur.eglise  # inchangé si non valide

        # Photo (optionnelle)
        if 'photo' in request.FILES:
            pasteur.photo = request.FILES['photo']

        pasteur.save()

        data = {
            'id': pasteur.id,
            'nom': pasteur.nom,
            'prenom': pasteur.prenom,
            'statut_actuel': pasteur.statut_actuel,
            'telephone': pasteur.telephone,
            'photo_url': pasteur.photo.url if pasteur.photo else '',
        }

        if is_ajax:
            return JsonResponse({'success': True, 'pasteur': data})
        messages.success(request, "Modification enregistrée avec succès.")
        return redirect('/personnel-pastoral/')
    except Exception as e:
        if is_ajax:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        messages.error(request, f"Erreur lors de la modification: {str(e)}")
        return redirect('/personnel-pastoral/')


def supprimer_pasteur(request, pasteur_id: int):
    """Supprime un pasteur. Retour JSON pour AJAX, redirection sinon."""
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if request.method != 'POST':
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'}, status=405)
        messages.error(request, "Méthode non autorisée pour la suppression.")
        return redirect('/personnel-pastoral/')

    try:
        pasteur = get_object_or_404(PersonnelPastoral, id=pasteur_id)
        nom_complet = f"{pasteur.prenom or ''} {pasteur.nom or ''}".strip()
        pasteur.delete()
        if is_ajax:
            return JsonResponse({'success': True, 'deleted_id': pasteur_id, 'message': f"Le pasteur {nom_complet or pasteur_id} a été supprimé avec succès."})
        messages.success(request, f"Le pasteur {nom_complet or pasteur_id} a été supprimé avec succès.")
    except Exception as e:
        if is_ajax:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        messages.error(request, f"Erreur lors de la suppression: {str(e)}")
    return redirect('/personnel-pastoral/')

@login_required
def les_membres(request):
    """Vue pour la page Les Membres"""
    membres = Membre.objects.all().order_by('nom', 'prenom')
    eglises = Eglise.objects.all().order_by('nom')
    pasteurs_count = PersonnelPastoral.objects.count()
    return render(request, 'main/les_membres.html', { 'membres': membres, 'eglises': eglises, 'pasteurs_count': pasteurs_count })

def membres_eglise(request):
    """Vue pour la page Membres d'Eglise"""
    membres = Membre.objects.all().order_by('nom', 'prenom')
    eglises = Eglise.objects.all().order_by('nom')
    pasteurs_count = PersonnelPastoral.objects.count()
    return render(request, 'main/membres_eglise.html', { 'membres': membres, 'eglises': eglises, 'pasteurs_count': pasteurs_count })

# =====================
# Finances: Edit/Delete
# =====================

@require_POST
def supprimer_finance(request, operation_id: int):
    """Supprime une opération Finance. Retour JSON (AJAX-friendly)."""
    try:
        op = get_object_or_404(Finance, id=operation_id)
        op.delete()
        return JsonResponse({'success': True, 'deleted_id': operation_id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@require_POST
def modifier_finance(request, operation_id: int):
    """Modifie une opération Finance avec un sous-ensemble de champs. Retour JSON."""
    try:
        op = get_object_or_404(Finance, id=operation_id)
        # Champs simples optionnels
        if 'libelle' in request.POST:
            op.libelle = request.POST.get('libelle') or op.libelle
        if 'description' in request.POST:
            op.description = request.POST.get('description') or ''
        if 'beneficiaire' in request.POST:
            op.beneficiaire = request.POST.get('beneficiaire') or ''
        if 'categorie' in request.POST:
            op.categorie = request.POST.get('categorie') or op.categorie
        if 'type_operation' in request.POST:
            t = request.POST.get('type_operation')
            if t in dict(Finance.TYPE_CHOICES):
                op.type_operation = t
        if 'montant' in request.POST:
            try:
                op.montant = Decimal(request.POST.get('montant'))
            except Exception:
                pass
        if 'date_operation' in request.POST:
            try:
                op.date_operation = datetime.strptime(request.POST.get('date_operation'), '%Y-%m-%d').date()
            except Exception:
                pass
        # Lier à un rapport (optionnel)
        if 'report_id' in request.POST:
            rid = request.POST.get('report_id')
            if rid:
                try:
                    op.report = FinanceReport.objects.get(id=int(rid))
                except Exception:
                    pass
            else:
                op.report = None
        op.save()
        return JsonResponse({
            'success': True,
            'operation': {
                'id': op.id,
                'eglise': op.eglise.nom,
                'report_id': op.report.id if op.report else None,
                'date_operation': op.date_operation.strftime('%Y-%m-%d'),
                'type_operation': op.type_operation,
                'libelle': op.libelle,
                'description': op.description,
                'beneficiaire': op.beneficiaire,
                'categorie': op.categorie,
                'montant': str(op.montant),
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@require_POST
def supprimer_finance_report(request, report_id: int):
    """Supprime un FinanceReport (et ses dépendances via CASCADE). Retour JSON."""
    try:
        rep = get_object_or_404(FinanceReport, id=report_id)
        rep.delete()
        return JsonResponse({'success': True, 'deleted_id': report_id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@require_POST
def modifier_finance_report(request, report_id: int):
    """Modifie un FinanceReport (en-tête, totaux). Retour JSON."""
    try:
        rep = get_object_or_404(FinanceReport, id=report_id)
        if 'responsable' in request.POST:
            rep.responsable = request.POST.get('responsable') or ''
        if 'contact' in request.POST:
            rep.contact = request.POST.get('contact') or ''
        if 'email' in request.POST:
            rep.email = request.POST.get('email') or ''
        if 'date_rapport' in request.POST:
            try:
                rep.date_rapport = datetime.strptime(request.POST.get('date_rapport'), '%Y-%m-%d').date()
            except Exception:
                pass
        if 'periode_du' in request.POST:
            try:
                rep.periode_du = datetime.strptime(request.POST.get('periode_du'), '%Y-%m-%d').date()
            except Exception:
                pass
        if 'periode_au' in request.POST:
            try:
                rep.periode_au = datetime.strptime(request.POST.get('periode_au'), '%Y-%m-%d').date()
            except Exception:
                pass
        if 'verifie_par' in request.POST:
            rep.verifie_par = request.POST.get('verifie_par') or ''
        if 'date_verification' in request.POST:
            try:
                rep.date_verification = datetime.strptime(request.POST.get('date_verification'), '%Y-%m-%d').date()
            except Exception:
                pass
        if 'approuve_par' in request.POST:
            rep.approuve_par = request.POST.get('approuve_par') or ''
        # Totaux (optionnels)
        for fld in ('total_recettes', 'total_depenses', 'solde_initial', 'solde_final'):
            if fld in request.POST:
                try:
                    setattr(rep, fld, Decimal(request.POST.get(fld)))
                except Exception:
                    pass
        rep.save()
        return JsonResponse({'success': True, 'report': {
            'id': rep.id,
            'eglise': rep.eglise.nom,
            'periode_du': rep.periode_du.strftime('%Y-%m-%d') if rep.periode_du else None,
            'periode_au': rep.periode_au.strftime('%Y-%m-%d') if rep.periode_au else None,
            'date_rapport': rep.date_rapport.strftime('%Y-%m-%d') if rep.date_rapport else None,
            'responsable': rep.responsable,
            'verifie_par': rep.verifie_par,
            'approuve_par': rep.approuve_par,
            'total_recettes': str(rep.total_recettes or 0),
            'total_depenses': str(rep.total_depenses or 0),
            'solde_initial': str(rep.solde_initial or 0),
            'solde_final': str(rep.solde_final or 0),
        }})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
def nos_fideles(request):
    """Vue pour la page Nos Fidèles (formulaire d'adhésion + tableau)."""
    eglises = Eglise.objects.all().order_by('nom')
    return render(request, 'main/nos_fideles.html', { 'eglises': eglises })


def comptabilite(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Créer un nouvel enregistrement comptable
            rapport = ComptabiliteReport.objects.create(
                eglise_nom=data.get('eglise_nom', ''),
                periode=data.get('periode', ''),
                responsable=data.get('responsable', ''),
                contact=data.get('contact', ''),
                email=data.get('email', ''),
                solde_initial=data.get('solde_initial', 0),
                total_revenus=data.get('total_revenus', 0),
                total_depenses=data.get('total_depenses', 0),
                solde_final=data.get('solde_final', 0),
                sources_revenus=data.get('sources_revenus', ''),
                dates_revenus=data.get('dates_revenus', ''),
                modes_paiement_revenus=data.get('modes_paiement_revenus', ''),
                evenements=data.get('evenements', ''),
                natures_depenses=data.get('natures_depenses', ''),
                montants_depenses=data.get('montants_depenses', ''),
                beneficiaires=data.get('beneficiaires', ''),
                justificatifs=data.get('justificatifs', ''),
                categories=data.get('categories', ''),
                nb_pj=data.get('nb_pj', 0),
                verifie_par=data.get('verifie_par', ''),
                approuve_par=data.get('approuve_par', ''),
                date_rapport=data.get('date_rapport', '')
            )
            
            return JsonResponse({'success': True, 'id': rapport.id})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    # GET request - afficher la page
    rapports = ComptabiliteReport.objects.all().order_by('-id')
    
    # Sérialiser les données pour JavaScript
    rapports_data = []
    for rapport in rapports:
        rapports_data.append({
            'id': rapport.id,
            'eglise_nom': rapport.eglise_nom,
            'periode': rapport.periode,
            'responsable': rapport.responsable,
            'contact': rapport.contact,
            'email': rapport.email,
            'solde_initial': rapport.solde_initial,
            'total_revenus': rapport.total_revenus,
            'total_depenses': rapport.total_depenses,
            'solde_final': rapport.solde_final,
            'sources_revenus': rapport.sources_revenus,
            'dates_revenus': rapport.dates_revenus,
            'modes_paiement_revenus': rapport.modes_paiement_revenus,
            'evenements': rapport.evenements,
            'natures_depenses': rapport.natures_depenses,
            'montants_depenses': rapport.montants_depenses,
            'beneficiaires': rapport.beneficiaires,
            'justificatifs': rapport.justificatifs,
            'categories': rapport.categories,
            'nb_pj': rapport.nb_pj,
            'verifie_par': rapport.verifie_par,
            'approuve_par': rapport.approuve_par,
            'date_rapport': rapport.date_rapport,
        })
    
    context = {
        'rapports_json': json.dumps(rapports_data, cls=DjangoJSONEncoder)
    }
    
    return render(request, 'main/comptabilite.html', context)

def supprimer_rapport_comptabilite(request, rapport_id):
    if request.method == 'DELETE':
        try:
            rapport = ComptabiliteReport.objects.get(id=rapport_id)
            rapport.delete()
            return JsonResponse({'success': True})
        except ComptabiliteReport.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Rapport non trouvé'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

@login_required
def banque(request):
    """Vue pour la page Banque (formulaire d'enregistrement financier de l'église + tableau)."""
    eglises = Eglise.objects.all().order_by('nom')
    
    if request.method == 'POST':
        try:
            is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
            # Récupération des données du formulaire
            eglise_id = request.POST.get('eglise_id')
            eglise = get_object_or_404(Eglise, id=eglise_id)
            
            # Création du rapport financier avec les nouvelles colonnes
            rapport = FinanceReport.objects.create(
                eglise=eglise,
                date_rapport=request.POST.get('date_rapport') or None,
                periode_du=request.POST.get('periode_debut') or None,
                periode_au=request.POST.get('periode_fin') or None,
                responsable=request.POST.get('responsable', ''),
                contact=request.POST.get('contact', ''),
                email=request.POST.get('email', ''),
                verifie_par=request.POST.get('verifie_par', ''),
                approuve_par=request.POST.get('approuve_par', ''),
                solde_initial=Decimal(request.POST.get('solde_initial', '0').replace(' ', '') or '0'),
                total_recettes=Decimal(request.POST.get('total_revenus', '0').replace(' ', '') or '0'),
                total_depenses=Decimal(request.POST.get('total_depenses', '0').replace(' ', '') or '0'),
                nombre_pieces_jointes=int((request.POST.get('nb_pj') or '0') or 0),
                # Nouvelles colonnes
                source_revenu=request.POST.get('source_revenu', ''),
                mode_paiement=request.POST.get('mode_paiement', ''),
                evenement=request.POST.get('evenement', ''),
                nature_depense=request.POST.get('nature_depense', ''),
                beneficiaire=request.POST.get('beneficiaire', ''),
                justificatif=request.POST.get('justificatif', ''),
            )
            
            # Calcul du solde final
            rapport.solde_final = rapport.solde_initial + rapport.total_recettes - rapport.total_depenses
            rapport.save()

            # Enregistrer les pièces jointes si fournies (globale et par ligne)
            try:
                saved_files_count = 0
                # Fichiers globaux
                for f in request.FILES.getlist('pj_global[]'):
                    FinanceAttachment.objects.create(report=rapport, nom=f.name, fichier=f)
                    saved_files_count += 1
                # Fichiers par ligne de dépense: clés de type depense_pj_<index>[]
                for key in list(request.FILES.keys()):
                    if key.startswith('depense_pj_'):
                        for f in request.FILES.getlist(key):
                            FinanceAttachment.objects.create(report=rapport, nom=f.name, fichier=f)
                            saved_files_count += 1
                # Mettre à jour le compteur si nécessaire
                if saved_files_count:
                    rapport.nombre_pieces_jointes = (rapport.nombre_pieces_jointes or 0) + saved_files_count
                    rapport.save(update_fields=['nombre_pieces_jointes'])
            except Exception:
                # Ne pas bloquer l'enregistrement du rapport si une PJ échoue
                pass
            
            if is_ajax:
                return JsonResponse({'success': True, 'id': rapport.id})
            else:
                messages.success(request, 'Rapport financier enregistré avec succès!')
                return redirect('main:banque')
            
        except Exception as e:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)}, status=400)
            messages.error(request, f"Erreur lors de l'enregistrement: {str(e)}")
    
    # Récupération des rapports existants (complets) et limitation à 10 pour l'affichage principal
    rapports_qs = FinanceReport.objects.select_related('eglise').all().order_by('-date_rapport', '-id')
    rapports = rapports_qs[:10]
    
    # Sérialiser les données pour JavaScript
    rapports_data = []
    for rapport in rapports:
        d = {
            'id': rapport.id,  # ID nécessaire pour la suppression
            'eglise__nom': rapport.eglise.nom,
            'periode_du': rapport.periode_du.strftime('%Y-%m-%d') if rapport.periode_du else '',
            'periode_au': rapport.periode_au.strftime('%Y-%m-%d') if rapport.periode_au else '',
            'responsable': rapport.responsable,
            'contact': rapport.contact,
            'email': rapport.email,
            'verifie_par': rapport.verifie_par,
            'approuve_par': rapport.approuve_par,
            'date_rapport': rapport.date_rapport.strftime('%Y-%m-%d') if rapport.date_rapport else '',
            'solde_initial': float(rapport.solde_initial or 0),
            'total_recettes': float(rapport.total_recettes or 0),
            'total_depenses': float(rapport.total_depenses or 0),
            'solde_final': float(rapport.solde_final or 0),
            'nombre_pieces_jointes': rapport.nombre_pieces_jointes or 0,
            'source_revenu': rapport.source_revenu,
            'mode_paiement': rapport.mode_paiement,
            'evenement': rapport.evenement,
            'nature_depense': rapport.nature_depense,
            'beneficiaire': rapport.beneficiaire,
            'justificatif': rapport.justificatif,
            # Champs détaillés si disponibles sur le modèle (compatibilité souple)
            'sources_revenus': getattr(rapport, 'sources_revenus', ''),
            'montants_revenus': getattr(rapport, 'montants_revenus', ''),
            'dates_revenus': getattr(rapport, 'dates_revenus', ''),
            'modes_paiement_revenus': getattr(rapport, 'modes_paiement_revenus', ''),
            'natures_depenses': getattr(rapport, 'natures_depenses', ''),
            'montants_depenses': getattr(rapport, 'montants_depenses', ''),
            'beneficiaires': getattr(rapport, 'beneficiaires', ''),
            'justificatifs': getattr(rapport, 'justificatifs', ''),
            'categories': getattr(rapport, 'categories', ''),
        }
        # Compléter avec agrégats si vides
        ag = _aggregate_finance_lines_for_report(rapport)
        for k in ('sources_revenus','montants_revenus','dates_revenus','modes_paiement_revenus','natures_depenses','montants_depenses','beneficiaires','justificatifs','categories'):
            if not d.get(k):
                d[k] = ag.get(k, '')
        # Pièces jointes
        d['attachments'] = _serialize_attachments_for_report(rapport)
        rapports_data.append(d)
    
    return render(request, 'main/banque.html', { 
        'eglises': eglises,
        'rapports': rapports,
        'rapports_json': json.dumps(rapports_data)
    })


@login_required
@ensure_csrf_cookie
def banque_records_all(request):
    """Affiche une page dédiée avec tous les rapports Banque (recherche, filtres, exports, pagination)."""
    rapports_qs = FinanceReport.objects.select_related('eglise').all().order_by('-date_rapport', '-id')

    rapports_data = []
    for rapport in rapports_qs:
        d = {
            'id': rapport.id,
            'eglise__nom': rapport.eglise.nom,
            'periode_du': rapport.periode_du.strftime('%Y-%m-%d') if rapport.periode_du else '',
            'periode_au': rapport.periode_au.strftime('%Y-%m-%d') if rapport.periode_au else '',
            'responsable': rapport.responsable,
            'contact': rapport.contact,
            'email': rapport.email,
            'verifie_par': rapport.verifie_par,
            'approuve_par': rapport.approuve_par,
            'date_rapport': rapport.date_rapport.strftime('%Y-%m-%d') if rapport.date_rapport else '',
            'solde_initial': float(rapport.solde_initial or 0),
            'total_recettes': float(rapport.total_recettes or 0),
            'total_depenses': float(rapport.total_depenses or 0),
            'solde_final': float(rapport.solde_final or 0),
            'nombre_pieces_jointes': rapport.nombre_pieces_jointes or 0,
            'source_revenu': rapport.source_revenu,
            'mode_paiement': rapport.mode_paiement,
            'evenement': rapport.evenement,
            'nature_depense': rapport.nature_depense,
            'beneficiaire': rapport.beneficiaire,
            'justificatif': rapport.justificatif,
            'sources_revenus': getattr(rapport, 'sources_revenus', ''),
            'montants_revenus': getattr(rapport, 'montants_revenus', ''),
            'dates_revenus': getattr(rapport, 'dates_revenus', ''),
            'modes_paiement_revenus': getattr(rapport, 'modes_paiement_revenus', ''),
            'natures_depenses': getattr(rapport, 'natures_depenses', ''),
            'montants_depenses': getattr(rapport, 'montants_depenses', ''),
            'beneficiaires': getattr(rapport, 'beneficiaires', ''),
            'justificatifs': getattr(rapport, 'justificatifs', ''),
            'categories': getattr(rapport, 'categories', ''),
        }
        ag = _aggregate_finance_lines_for_report(rapport)
        for k in ('sources_revenus','montants_revenus','dates_revenus','modes_paiement_revenus','natures_depenses','montants_depenses','beneficiaires','justificatifs','categories'):
            if not d.get(k):
                d[k] = ag.get(k, '')
        d['attachments'] = _serialize_attachments_for_report(rapport)
        rapports_data.append(d)

    return render(request, 'main/banque_records_all.html', {
        'rapports_json': json.dumps(rapports_data)
    })


def supprimer_rapport_banque(request, rapport_id):
    """Supprimer définitivement un rapport financier de la base de données"""
    if request.method == 'POST':
        try:
            rapport = get_object_or_404(FinanceReport, id=rapport_id)
            eglise_nom = rapport.eglise.nom
            rapport.delete()
            
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True, 
                    'message': f'Rapport de {eglise_nom} supprimé définitivement'
                })
            else:
                messages.success(request, f'Rapport de {eglise_nom} supprimé définitivement')
                return redirect('main:banque')
                
        except Exception as e:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False, 
                    'error': f'Erreur lors de la suppression: {str(e)}'
                })
            else:
                messages.error(request, f'Erreur lors de la suppression: {str(e)}')
                return redirect('main:banque')
    
    return redirect('main:banque')


def ajouter_membre(request):
    """Crée un membre d'église à partir du formulaire d'adhésion (modal)."""
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if request.method != 'POST':
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'}, status=405)
        return redirect('main:membres_eglise')

    try:
        # Récupération champs formulaire (modal)
        nom_complet = (request.POST.get('nom_complet') or '').strip()
        # Split simple: dernier mot = nom, le reste = prenom
        nom_parts = nom_complet.split()
        nom = nom_parts[-1] if nom_parts else ''
        prenom = ' '.join(nom_parts[:-1]) if len(nom_parts) > 1 else ''

        # Normalisations et récupérations robustes
        raw_sexe = (request.POST.get('sexe') or '').strip()
        # Supporte "M", "F", "Masculin", "Féminin"
        if raw_sexe:
            sx = raw_sexe.strip().upper()
            if sx.startswith('M'):
                sexe = 'M'
            elif sx.startswith('F'):
                sexe = 'F'
            else:
                sexe = sx  # on laisse tel quel, validation plus bas
        else:
            sexe = ''

        date_naissance = request.POST.get('date_naissance') or None
        lieu_naissance = request.POST.get('lieu_naissance') or ''
        adresse = request.POST.get('adresse') or ''
        telephone = request.POST.get('telephone') or ''
        email = request.POST.get('email') or ''
        profession = request.POST.get('profession') or ''
        etat_civil = request.POST.get('situation_familiale') or request.POST.get('etat_civil') or ''
        nombre_enfants = request.POST.get('nombre_enfants') or '0'
        eglise_id = request.POST.get('eglise_id')
        nom_eglise = (request.POST.get('nom_eglise') or '').strip()

        # Baptême
        def to_bool(val: str) -> bool:
            v = (val or '').strip().lower()
            return v in ('oui', 'true', '1', 'on', 'yes')

        baptise = to_bool(request.POST.get('baptise')) or to_bool(request.POST.get('est_baptise'))
        # Accept both keys
        date_bapteme = request.POST.get('date_bapteme') or request.POST.get('date_conversion') or None
        lieu_bapteme = request.POST.get('lieu_bapteme') or ''

        # Champs étendus
        motivation = request.POST.get('motivation', '')
        services_str = request.POST.get('services', '')  # CSV depuis le formulaire
        soutien_financier_raw = request.POST.get('soutien_financier', '')
        montant_souhaite_raw = request.POST.get('montant_souhaite', '')
        # Départements (CSV ou texte libre)
        departement_str = request.POST.get('departement', '')

        # Date d'adhésion (facultative dans le modal): défaut = aujourd'hui
        date_adhesion = request.POST.get('date_adhesion') or datetime.today().strftime('%Y-%m-%d')

        # Validations minimales
        if not nom or not prenom:
            if is_ajax:
                return JsonResponse({'success': False, 'error': "Veuillez saisir le nom complet (Prénom(s) et Nom)."}, status=400)
            messages.error(request, "Veuillez saisir le nom complet (Prénom(s) et Nom).")
            return redirect('main:membres_eglise')
        if sexe not in ('M', 'F'):
            if is_ajax:
                return JsonResponse({'success': False, 'error': "Veuillez sélectionner le sexe (Masculin ou Féminin)."}, status=400)
            messages.error(request, "Veuillez sélectionner le sexe (Masculin ou Féminin).")
            return redirect('main:membres_eglise')
        # Église: accepte eglise_id OU nom_eglise
        eglise = None
        if eglise_id:
            eglise = get_object_or_404(Eglise, id=eglise_id)
        elif nom_eglise:
            eglise = Eglise.objects.filter(nom__iexact=nom_eglise).first() or Eglise.objects.filter(nom__icontains=nom_eglise).first()
            if not eglise:
                if is_ajax:
                    return JsonResponse({'success': False, 'error': "Église introuvable. Veuillez sélectionner une église valide."}, status=400)
                messages.error(request, "Église introuvable. Veuillez sélectionner une église valide.")
                return redirect('main:membres_eglise')
        else:
            if is_ajax:
                return JsonResponse({'success': False, 'error': "Veuillez sélectionner l'Église d'appartenance."}, status=400)
            messages.error(request, "Veuillez sélectionner l'Église d'appartenance.")
            return redirect('main:membres_eglise')

        # Aides de parsing
        def parse_date(val):
            if not val:
                return None
            try:
                return datetime.strptime(val, '%Y-%m-%d').date()
            except Exception:
                return None
        def parse_int(val, default=0):
            try:
                return int(val)
            except Exception:
                return default
        def parse_decimal(val):
            try:
                from decimal import Decimal
                return Decimal(val)
            except Exception:
                return None

        membre = Membre(
            eglise=eglise,
            nom=nom,
            prenom=prenom,
            sexe=sexe,
            date_naissance=parse_date(date_naissance),
            lieu_naissance=lieu_naissance,
            adresse=adresse,
            telephone=telephone,
            email=email,
            profession=profession,
            date_bapteme=parse_date(date_bapteme),
            lieu_bapteme=lieu_bapteme,
            date_adhesion=parse_date(date_adhesion) or datetime.today().date(),
            etat_civil=etat_civil or 'celibataire',
            nombre_enfants=parse_int(nombre_enfants, 0),
            motivation=motivation,
            services=services_str,
            departement=departement_str,
            soutien_financier=to_bool(soutien_financier_raw),
            montant_souhaite=parse_decimal(montant_souhaite_raw),
        )
        # Pièces jointes si fournies
        if 'certificat_bapteme' in request.FILES:
            membre.certificat_bapteme = request.FILES['certificat_bapteme']
        # Champs optionnels absents du model: ignorer (baptisé par, etc.)
        membre.save()

        if is_ajax:
            # Obtenir tailles de fichiers si présents
            def fsize(f):
                try:
                    return f.size if f else None
                except Exception:
                    return None

            data = {
                'success': True,
                'membre': {
                    'id': membre.id,
                    'nom': membre.nom,
                    'prenom': membre.prenom,
                    'sexe': membre.sexe,
                    'date_naissance': membre.date_naissance.isoformat() if membre.date_naissance else None,
                    'lieu_naissance': membre.lieu_naissance,
                    'adresse': membre.adresse,
                    'telephone': membre.telephone,
                    'email': membre.email,
                    'profession': membre.profession,
                    'etat_civil': membre.etat_civil,
                    'nombre_enfants': membre.nombre_enfants,
                    'date_bapteme': membre.date_bapteme.isoformat() if membre.date_bapteme else None,
                    'lieu_bapteme': membre.lieu_bapteme,
                    'date_adhesion': membre.date_adhesion.isoformat() if membre.date_adhesion else None,
                    'motivation': membre.motivation,
                    'services': membre.services,
                    'departement': membre.departement,
                    'soutien_financier': membre.soutien_financier,
                    'montant_souhaite': str(membre.montant_souhaite) if membre.montant_souhaite is not None else None,
                    # Documents
                    'certificat_bapteme_url': membre.certificat_bapteme.url if getattr(membre, 'certificat_bapteme', None) else '',
                    'certificat_bapteme_size': fsize(getattr(membre, 'certificat_bapteme', None)),
                    'eglise': {
                        'id': membre.eglise.id if membre.eglise else None,
                        'nom': membre.eglise.nom if membre.eglise else None,
                    }
                }
            }
            return JsonResponse(data)
        else:
            messages.success(request, "Membre ajouté avec succès.")
            return redirect('main:membres_eglise')
    except Exception as e:
        if is_ajax:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
        messages.error(request, "Erreur lors de l'ajout du membre.")
        return redirect('main:membres_eglise')

def supprimer_membre(request, membre_id: int):
    """Supprime un membre. Retour JSON pour AJAX, redirection sinon."""
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if request.method != 'POST':
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'}, status=405)
        messages.error(request, "Méthode non autorisée pour la suppression.")
        return redirect('main:membres_eglise')

    try:
        membre = get_object_or_404(Membre, id=membre_id)
        nom_complet = f"{membre.prenom or ''} {membre.nom or ''}".strip()
        membre.delete()
        if is_ajax:
            return JsonResponse({'success': True, 'deleted_id': membre_id, 'message': f"Le membre {nom_complet or membre_id} a été supprimé avec succès."})
        messages.success(request, f"Le membre {nom_complet or membre_id} a été supprimé avec succès.")
    except Exception as e:
        if is_ajax:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        messages.error(request, f"Erreur lors de la suppression: {str(e)}")
    return redirect('main:membres_eglise')

@require_POST
def supprimer_departement(request, departement_id: int):
    """Supprime un département. Retour JSON pour AJAX, redirection sinon.

    - Protégé par CSRF (POST requis)
    - Si en-tête `X-Requested-With: XMLHttpRequest` présent, renvoie JSON
    - Sinon, utilise messages + redirection vers la page des départements
    """
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    try:
        dep = get_object_or_404(Departement, id=departement_id)
        nom_dep = dep.nom or str(departement_id)
        # Supprimer d'abord les pièces jointes si présentes (évite contraintes FK PROTECT)
        try:
            for att in list(getattr(dep, 'attachments', []).all()):
                try:
                    if getattr(att, 'fichier', None):
                        att.fichier.delete(save=False)
                except Exception:
                    # Ignorer les erreurs de stockage
                    pass
                try:
                    att.delete()
                except Exception:
                    pass
        except Exception:
            pass
        # Supprimer le département
        dep.delete()
        if is_ajax:
            return JsonResponse({'success': True, 'deleted_id': departement_id, 'message': f"Le département {nom_dep} a été supprimé avec succès."})
        messages.success(request, f"Le département {nom_dep} a été supprimé avec succès.")
    except Exception as e:
        if is_ajax:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        messages.error(request, f"Erreur lors de la suppression du département: {str(e)}")
    return redirect('main:departements')

def details_membre(request, membre_id: int):
    """Affiche les détails d'un membre."""
    membre = get_object_or_404(Membre, id=membre_id)
    return render(request, 'main/membre_details.html', {'membre': membre})

# Enregistrement des effectifs du culte (API JSON)
@require_POST
def enregistrer_effectif_culte(request):
    try:
        data = {}
        if request.headers.get('Content-Type', '').startswith('application/json'):
            data = json.loads(request.body.decode('utf-8') or '{}')
        else:
            # fallback formulaire classique
            data = request.POST.dict()

        def to_int(val, default=0):
            try:
                return max(0, int(str(val).strip()))
            except Exception:
                return default

        date_str = (data.get('date') or '').strip()
        if not date_str:
            return JsonResponse({'success': False, 'error': "La date est obligatoire."}, status=400)
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            return JsonResponse({'success': False, 'error': "Format de date invalide (AAAA-MM-JJ)."}, status=400)

        eglise_id = data.get('eglise_id') or data.get('eglise') or ''
        if not str(eglise_id).strip():
            return JsonResponse({'success': False, 'error': "Veuillez sélectionner une Église."}, status=400)
        try:
            eglise = Eglise.objects.get(id=int(eglise_id))
        except Exception:
            return JsonResponse({'success': False, 'error': "Église introuvable."}, status=400)

        hommes = to_int(data.get('hommes'))
        femmes = to_int(data.get('femmes'))
        filles = to_int(data.get('filles'))
        garcons = to_int(data.get('garcons'))
        nouveaux = to_int(data.get('nouveaux'))
        total_enfants = filles + garcons
        total = hommes + femmes + total_enfants + nouveaux

        # Upsert par date (prévenir doublons)
        eff, created = CulteEffectif.objects.update_or_create(
            date=d,
            eglise=eglise,
            defaults={
                'hommes': hommes,
                'femmes': femmes,
                'filles': filles,
                'garcons': garcons,
                'nouveaux': nouveaux,
                'total_enfants': total_enfants,
                'total': total,
            }
        )
        return JsonResponse({'success': True, 'id': eff.id, 'total': eff.total, 'date': eff.date.strftime('%Y-%m-%d'), 'updated': (not created)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@require_GET
def lister_effectifs_culte(request):
    try:
        try:
            limit = int(request.GET.get('limit', '10'))
        except Exception:
            limit = 10
        limit = max(1, min(limit, 100))
        qs = CulteEffectif.objects.all()

        eglise_id = request.GET.get('eglise_id')
        if eglise_id:
            try:
                qs = qs.filter(eglise_id=int(eglise_id))
            except Exception:
                pass
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        from datetime import datetime as _dt
        if date_from:
            try:
                df = _dt.strptime(date_from, '%Y-%m-%d').date()
                qs = qs.filter(date__gte=df)
            except Exception:
                pass
        if date_to:
            try:
                dt = _dt.strptime(date_to, '%Y-%m-%d').date()
                qs = qs.filter(date__lte=dt)
            except Exception:
                pass

        # Tri et limitation (appliquer l'ordre avant la coupe). Si all=1, ne pas limiter
        qs = qs.order_by('-date', '-id')
        show_all = (request.GET.get('all') in ('1', 'true', 'True', 'yes', 'oui'))
        if not show_all:
            qs = qs[:limit]
        data = [
            {
                'id': e.id,
                'date': e.date.strftime('%Y-%m-%d'),
                'eglise': getattr(e.eglise, 'nom', None),
                'eglise_id': getattr(e.eglise, 'id', None),
                'hommes': e.hommes,
                'femmes': e.femmes,
                'filles': e.filles,
                'garcons': e.garcons,
                'total_enfants': e.total_enfants,
                'nouveaux': e.nouveaux,
                'total': e.total,
            }
            for e in qs
        ]
        return JsonResponse({'success': True, 'items': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

import csv
from django.http import HttpResponse
from django.utils.text import slugify
from io import BytesIO
try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
except Exception:
    # Le module peut ne pas être installé; la vue gérera ce cas proprement
    A4 = None

def exporter_effectifs_culte_csv(request):
    try:
        qs = CulteEffectif.objects.all().order_by('-date', '-id')
        eglise_id = request.GET.get('eglise_id')
        eglise_nom = ''
        if eglise_id:
            try:
                eglise_obj = Eglise.objects.get(id=int(eglise_id))
                eglise_nom = eglise_obj.nom
                qs = qs.filter(eglise_id=eglise_obj.id)
            except Exception:
                pass
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        from datetime import datetime as _dt
        if date_from:
            try:
                df = _dt.strptime(date_from, '%Y-%m-%d').date()
                qs = qs.filter(date__gte=df)
            except Exception:
                pass
        if date_to:
            try:
                dt = _dt.strptime(date_to, '%Y-%m-%d').date()
                qs = qs.filter(date__lte=dt)
            except Exception:
                pass

        # Construire un nom de fichier explicite avec église et période
        eg_slug = slugify(eglise_nom) if eglise_nom else 'toutes-eglises'
        df_part = (date_from or 'toutes').replace(':', '-')
        dt_part = (date_to or 'toutes').replace(':', '-')
        filename = f"effectifs_culte_{eg_slug}_{df_part}_{dt_part}.csv"

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)

        # Évaluer la queryset pour calculer les totaux et le nombre de lignes
        # Optionnel: limiter le nombre de lignes exportées
        limit = request.GET.get('limit')
        if limit:
            try:
                qs = qs[:max(0, int(limit))]
            except Exception:
                pass

        items = list(qs)
        nb = len(items)
        s_h, s_f, s_filles, s_garcons, s_te, s_nv, s_tot = 0, 0, 0, 0, 0, 0, 0
        for e in items:
            s_h += e.hommes or 0
            s_f += e.femmes or 0
            s_filles += e.filles or 0
            s_garcons += e.garcons or 0
            s_te += e.total_enfants or 0
            s_nv += e.nouveaux or 0
            s_tot += e.total or 0

        # Indicateur de nombre d'enregistrements
        writer.writerow(["Nombre d'enregistrements", nb])
        writer.writerow([])

        # En-têtes + données
        writer.writerow(['Date', 'Eglise', 'Hommes', 'Femmes', 'Filles', 'Garcons', 'Total enfants', 'Nouveaux', 'Total'])
        for e in items:
            _eglise_nom = getattr(e.eglise, 'nom', '')
            if _eglise_nom is None:
                _eglise_nom = ''
            _eglise_nom = str(_eglise_nom).replace('\r', ' ').replace('\n', ' ').strip()
            _eglise_nom = _ud.normalize('NFC', _eglise_nom)
            _eglise_nom = _fix_mojibake(_eglise_nom)
            writer.writerow([
                e.date.strftime('%Y-%m-%d'),
                _eglise_nom,
                e.hommes,
                e.femmes,
                e.filles,
                e.garcons,
                e.total_enfants,
                e.nouveaux,
                e.total,
            ])

        # Ligne de totaux
        writer.writerow(['TOTAL', '', s_h, s_f, s_filles, s_garcons, s_te, s_nv, s_tot])
        return response
    except Exception as e:
        return HttpResponse(f"Erreur export: {e}", status=500)

def exporter_effectifs_culte_excel(request):
    """Export Excel (compatibilité) en servant un CSV sous mimetype Excel.

    Filtres acceptés: eglise_id, date_from, date_to, limit (facultatif)
    """
    try:
        qs = CulteEffectif.objects.all().order_by('-date', '-id')
        eglise_id = request.GET.get('eglise_id')
        eglise_nom = ''
        if eglise_id:
            try:
                eglise_obj = Eglise.objects.get(id=int(eglise_id))
                eglise_nom = eglise_obj.nom
                qs = qs.filter(eglise_id=eglise_obj.id)
            except Exception:
                pass
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        from datetime import datetime as _dt
        if date_from:
            try:
                df = _dt.strptime(date_from, '%Y-%m-%d').date()
                qs = qs.filter(date__gte=df)
            except Exception:
                pass
        if date_to:
            try:
                dt = _dt.strptime(date_to, '%Y-%m-%d').date()
                qs = qs.filter(date__lte=dt)
            except Exception:
                pass

        # Limitation optionnelle
        limit = request.GET.get('limit')
        if limit:
            try:
                qs = qs[:max(0, int(limit))]
            except Exception:
                pass

        items = list(qs)
        eg_slug = slugify(eglise_nom) if eglise_nom else 'toutes-eglises'
        df_part = (date_from or 'toutes').replace(':', '-')
        dt_part = (date_to or 'toutes').replace(':', '-')
        filename = f"effectifs_culte_{eg_slug}_{df_part}_{dt_part}.csv"

        # Réponse Excel-friendly (CSV UTF-8 avec BOM pour accents)
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        # BOM UTF-8 pour Excel (accents)
        try:
            response.write('\ufeff')
        except Exception:
            pass
        # Indication du séparateur pour Excel
        try:
            response.write('sep=;\r\n')
        except Exception:
            pass

        writer = csv.writer(response, delimiter=';', lineterminator='\r\n', quoting=csv.QUOTE_MINIMAL)
        # En-têtes (ASCII pour compatibilité ouverture directe dans Excel)
        writer.writerow(['Date', 'Eglise', 'Hommes', 'Femmes', 'Filles', 'Garçons', 'Total enfants', 'Nouveaux', 'Total'])

        s_h = s_f = s_filles = s_garcons = s_te = s_nv = s_tot = 0
        for e in items:
            # Assainir le nom d'église (pas de retours ligne)
            _eglise_nom = getattr(e.eglise, 'nom', '')
            if _eglise_nom is None:
                _eglise_nom = ''
            _eglise_nom = str(_eglise_nom).replace('\r', ' ').replace('\n', ' ').strip()
            writer.writerow([
                e.date.strftime('%Y-%m-%d'),
                _eglise_nom,
                e.hommes,
                e.femmes,
                e.filles,
                e.garcons,
                e.total_enfants,
                e.nouveaux,
                e.total,
            ])
            s_h += e.hommes or 0
            s_f += e.femmes or 0
            s_filles += e.filles or 0
            s_garcons += e.garcons or 0
            s_te += e.total_enfants or 0
            s_nv += e.nouveaux or 0
            s_tot += e.total or 0
        # Totaux
        writer.writerow(['TOTAL', '', s_h, s_f, s_filles, s_garcons, s_te, s_nv, s_tot])

        return response
    except Exception as e:
        return HttpResponse(f"Erreur export Excel: {e}", status=500)

def exporter_effectifs_culte_xlsx(request):
    """Export natif Excel .xlsx (accents fiables). Fallback CSV si openpyxl indisponible.

    Filtres acceptés: eglise_id, date_from, date_to, limit (facultatif)
    """
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except Exception:
            # Fallback vers CSV Excel si openpyxl absent
            return exporter_effectifs_culte_excel(request)

        qs = CulteEffectif.objects.all().order_by('-date', '-id')
        eglise_id = request.GET.get('eglise_id')
        eglise_nom = ''
        if eglise_id:
            try:
                eglise_obj = Eglise.objects.get(id=int(eglise_id))
                eglise_nom = eglise_obj.nom or ''
                qs = qs.filter(eglise_id=eglise_obj.id)
            except Exception:
                pass
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        from datetime import datetime as _dt
        if date_from:
            try:
                df = _dt.strptime(date_from, '%Y-%m-%d').date()
                qs = qs.filter(date__gte=df)
            except Exception:
                pass
        if date_to:
            try:
                dt = _dt.strptime(date_to, '%Y-%m-%d').date()
                qs = qs.filter(date__lte=dt)
            except Exception:
                pass

        limit = request.GET.get('limit')
        if limit:
            try:
                qs = qs[:max(0, int(limit))]
            except Exception:
                pass

        items = list(qs)
        eg_slug = slugify(eglise_nom) if eglise_nom else 'toutes-eglises'
        df_part = (date_from or 'toutes').replace(':', '-')
        dt_part = (date_to or 'toutes').replace(':', '-')
        filename = f"effectifs_culte_{eg_slug}_{df_part}_{dt_part}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = 'Effectifs'

        headers = ['Date', 'Église', 'Hommes', 'Femmes', 'Filles', 'Garçons', 'Total enfants', 'Nouveaux', 'Total']
        ws.append(headers)
        # Style en-tête
        head_font = Font(bold=True)
        for col in range(1, len(headers)+1):
            c = ws.cell(row=1, column=col)
            c.font = head_font
            c.alignment = Alignment(vertical='center')

        s_h = s_f = s_filles = s_garcons = s_te = s_nv = s_tot = 0
        for e in items:
            egname = getattr(e.eglise, 'nom', '') or ''
            egname = str(egname).replace('\r',' ').replace('\n',' ').strip()
            egname = _ud.normalize('NFC', egname)
            egname = _fix_mojibake(egname)
            row = [
                e.date.strftime('%Y-%m-%d'),
                egname,
                e.hommes,
                e.femmes,
                e.filles,
                e.garcons,
                e.total_enfants,
                e.nouveaux,
                e.total,
            ]
            ws.append(row)
            s_h += e.hommes or 0
            s_f += e.femmes or 0
            s_filles += e.filles or 0
            s_garcons += e.garcons or 0
            s_te += e.total_enfants or 0
            s_nv += e.nouveaux or 0
            s_tot += e.total or 0

        # Totaux
        ws.append(['TOTAL', '', s_h, s_f, s_filles, s_garcons, s_te, s_nv, s_tot])
        last_row = ws.max_row
        ws.cell(row=last_row, column=1).font = Font(bold=True)

        # Largeurs de colonnes automatiques approximatives
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    val = str(cell.value) if cell.value is not None else ''
                    if len(val) > max_len:
                        max_len = len(val)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(40, max(10, max_len + 2))

        from io import BytesIO
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        response = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        return HttpResponse(f"Erreur export XLSX: {e}", status=500)

@require_http_methods(["GET"])
def api_sp_export_pdf(request):
    """Export PDF des enregistrements StrategicPlan via ReportLab.

    Même logique de filtrage/tri que CSV/XLSX.
    Retourne 501 si ReportLab n'est pas disponible.
    """
    if A4 is None:
        return HttpResponse(
            "ReportLab n'est pas installé sur le serveur. Veuillez installer 'reportlab' ou utiliser l'export Excel.",
            status=501
        )
    try:
        qs = StrategicPlan.objects.select_related('eglise').all().order_by('-date_prev','heure','activite')
        eg = request.GET.get('eglise_id')
        if eg:
            try:
                qs = qs.filter(eglise_id=int(eg))
            except Exception:
                pass
        # Filtres de période (optionnels, comme dans api_sp_list)
        df = (request.GET.get('date_from') or '').strip()
        dt = (request.GET.get('date_to') or '').strip()
        if df:
            try:
                from datetime import datetime as _dt
                dfrom = _dt.strptime(df, '%Y-%m-%d').date()
                qs = qs.filter(date_prev__gte=dfrom)
            except Exception:
                pass
        if dt:
            try:
                from datetime import datetime as _dt
                dto = _dt.strptime(dt, '%Y-%m-%d').date()
                qs = qs.filter(date_prev__lte=dto)
            except Exception:
                pass

        # Préparation des données
        headers = [
            'Église','Département','Type','Activité/Événement','Jour','Date','Heure','Lieu',
            'Besoins','Responsable','Objectif','Budget','Statut','Validé par','Date validation'
        ]
        rows = []
        for r in qs:
            eglise_nom = _fix_mojibake(_to_str(getattr(getattr(r,'eglise',None),'nom','')))
            departement = _fix_mojibake(_to_str(r.departement))
            type_plan = _fix_mojibake(_to_str(getattr(r, 'type_planification', '')))
            activite = _fix_mojibake(_to_str(r.activite))
            jour = _fix_mojibake(_to_str(r.jour))
            date_prev = r.date_prev.strftime('%Y-%m-%d') if r.date_prev else ''
            heure = r.heure.strftime('%H:%M') if r.heure else ''
            lieu = _fix_mojibake(_to_str(r.lieu))
            besoins = _fix_mojibake(_to_str(r.besoins))
            responsable = _fix_mojibake(_to_str(r.responsable))
            objectif = _fix_mojibake(_to_str(r.objectif))
            budget = str(r.budget) if r.budget is not None else ''
            statut = _fix_mojibake(_to_str(r.statut))
            nom_resp = _fix_mojibake(_to_str(r.nom_resp))
            nom_pasteur = _fix_mojibake(_to_str(r.nom_pasteur))
            date_valid = r.date_valid.strftime('%Y-%m-%d') if r.date_valid else ''
            valide_par = (nom_resp + ' / ' + nom_pasteur).strip(' / ')
            rows.append([
                eglise_nom, departement, type_plan, activite, jour, date_prev, heure, lieu,
                besoins, responsable, objectif, budget, statut, valide_par, date_valid
            ])

        # Construction PDF améliorée (paysage, en-tête/pied, wrapping, colonnes ajustées)
        from io import BytesIO as _BIO
        buffer = _BIO()

        # Mise en page paysage pour cas à nombreuses colonnes
        pagesize = landscape(A4)
        doc = SimpleDocTemplate(
            buffer, pagesize=pagesize,
            rightMargin=24, leftMargin=24, topMargin=36, bottomMargin=30
        )
        styles = getSampleStyleSheet()
        style_title = styles['Title']
        style_cell = styles['Normal']
        style_cell.fontSize = 8
        style_cell.leading = 10

        # En-tête/pied de page
        from datetime import datetime as _now
        def _on_page(canvas, _doc):
            canvas.saveState()
            canvas.setFont('Helvetica-Bold', 10)
            canvas.drawString(_doc.leftMargin, _doc.height + _doc.topMargin - 10, 'Planification Stratégique')
            canvas.setFont('Helvetica', 8)
            ts = _now.now().strftime('%Y-%m-%d %H:%M')
            canvas.drawRightString(_doc.width + _doc.leftMargin, _doc.height + _doc.topMargin - 10, f'Généré le {ts}')
            # Pied de page: pagination
            canvas.setFont('Helvetica', 8)
            canvas.drawRightString(_doc.width + _doc.leftMargin, 15, f'Page {canvas.getPageNumber()}')
            canvas.restoreState()

        # Conversion Paragraph pour wrapping
        data = [headers]
        def P(txt):
            return Paragraph(_fix_mojibake(_to_str(txt)), style_cell)
        for r in rows:
            data.append([P(x) for x in r])

        # Largeurs de colonnes approximatives, adaptées à la largeur disponible
        # [eglise, departement, type, activite, jour, date, heure, lieu, besoins, responsable, objectif, budget, statut, valide_par, date_valid]
        base_widths = [80, 80, 70, 120, 60, 70, 45, 90, 120, 90, 120, 60, 70, 110, 70]
        total_base = sum(base_widths)
        avail = doc.width
        col_widths = [max(35, (w/total_base)*avail) for w in base_widths]

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4e73df')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#dddddd')),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.white]),
        ]))

        story = [Spacer(1, 6), table]
        doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
        pdf = buffer.getvalue()
        buffer.close()

        resp = HttpResponse(content_type='application/pdf')
        resp['Content-Disposition'] = 'attachment; filename="planification_strategique.pdf"'
        resp.write(pdf)
        return resp
    except Exception as e:
        return HttpResponse(f"Erreur export PDF: {e}", status=500)

@login_required
@ensure_csrf_cookie
def planification_complet(request):
    """Vue pour la page Planning complet (tous les enregistrements + filtres + exports)."""
    return render(request, 'main/planification_full.html')

def exporter_effectifs_culte_pdf(request):
    # Vérifie disponibilité de reportlab
    if A4 is None:
        return HttpResponse(
            "ReportLab n'est pas installé sur le serveur. Veuillez installer 'reportlab' ou utiliser l'export CSV.",
            status=501
        )
    try:
        qs = CulteEffectif.objects.all().order_by('-date', '-id')
        eglise_id = request.GET.get('eglise_id')
        eglise_nom = ''
        if eglise_id:
            try:
                eglise_obj = Eglise.objects.get(id=int(eglise_id))
                eglise_nom = eglise_obj.nom
                qs = qs.filter(eglise_id=eglise_obj.id)
            except Exception:
                pass
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        from datetime import datetime as _dt
        if date_from:
            try:
                df = _dt.strptime(date_from, '%Y-%m-%d').date()
                qs = qs.filter(date__gte=df)
            except Exception:
                pass
        if date_to:
            try:
                dt = _dt.strptime(date_to, '%Y-%m-%d').date()
                qs = qs.filter(date__lte=dt)
            except Exception:
                pass

        items = list(qs)

        # Nom de fichier dynamique
        eg_slug = slugify(eglise_nom) if eglise_nom else 'toutes-eglises'
        df_part = (date_from or 'toutes').replace(':', '-')
        dt_part = (date_to or 'toutes').replace(':', '-')
        filename = f"effectifs_culte_{eg_slug}_{df_part}_{dt_part}.pdf"

        # Génération PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, title="Effectifs du culte")
        styles = getSampleStyleSheet()
        story = []

        titre = Paragraph("Effectifs du culte", styles['Title'])
        story.append(titre)
        meta_txt = f"Église: <b>{eglise_nom or 'Toutes les églises'}</b> — Période: <b>{date_from or 'Toutes'}</b> → <b>{date_to or 'Toutes'}</b>"
        story.append(Paragraph(meta_txt, styles['Normal']))
        story.append(Spacer(1, 10))

        # Tableau
        data = [[
            'Date', 'Église', 'Hommes', 'Femmes', 'Filles', 'Garçons', 'Total enfants', 'Nouveaux', 'Total'
        ]]
        s_h = s_f = s_fi = s_ga = s_te = s_nv = s_tot = 0
        for e in items:
            data.append([
                e.date.strftime('%Y-%m-%d'),
                getattr(e.eglise, 'nom', ''),
                e.hommes,
                e.femmes,
                e.filles,
                e.garcons,
                e.total_enfants,
                e.nouveaux,
                e.total,
            ])
            s_h += e.hommes or 0
            s_f += e.femmes or 0
            s_fi += e.filles or 0
            s_ga += e.garcons or 0
            s_te += e.total_enfants or 0
            s_nv += e.nouveaux or 0
            s_tot += e.total or 0

        data.append([
            'TOTAL', '', s_h, s_f, s_fi, s_ga, s_te, s_nv, s_tot
        ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f5f5f5')),
            ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#cccccc')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (2,1), (-1,-1), 'RIGHT'),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#fafafa')),
        ]))
        story.append(table)

        doc.build(story)
        pdf = buffer.getvalue()
        buffer.close()

        resp = HttpResponse(content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        resp.write(pdf)
        return resp
    except Exception as e:
        return HttpResponse(f"Erreur export PDF: {e}", status=500)


@require_POST
def supprimer_effectif_culte(request, eff_id: int):
    try:
        eff = get_object_or_404(CulteEffectif, id=eff_id)
        eff.delete()
        return JsonResponse({'success': True, 'deleted_id': eff_id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@require_POST
def modifier_effectif_culte(request, eff_id: int):
    try:
        eff = get_object_or_404(CulteEffectif, id=eff_id)
        if request.headers.get('Content-Type', '').startswith('application/json'):
            data = json.loads(request.body.decode('utf-8') or '{}')
        else:
            data = request.POST.dict()

        def to_int(val, default=0):
            try:
                return max(0, int(str(val).strip()))
            except Exception:
                return default

        # Facultatif: mise à jour église/date
        eglise_id = data.get('eglise_id') or data.get('eglise')
        if eglise_id is not None:
            eglise = get_object_or_404(Eglise, id=int(eglise_id))
            eff.eglise = eglise

        date_str = data.get('date')
        if date_str is not None:
            try:
                d = datetime.strptime(date_str, '%Y-%m-%d').date()
            except Exception:
                return JsonResponse({'success': False, 'error': "Format de date invalide (AAAA-MM-JJ)."}, status=400)
            eff.date = d

        # Champs numériques
        fields = {
            'hommes': to_int(data.get('hommes'), eff.hommes),
            'femmes': to_int(data.get('femmes'), eff.femmes),
            'filles': to_int(data.get('filles'), eff.filles),
            'garcons': to_int(data.get('garcons'), eff.garcons),
            'nouveaux': to_int(data.get('nouveaux'), eff.nouveaux),
        }
        eff.hommes = fields['hommes']
        eff.femmes = fields['femmes']
        eff.filles = fields['filles']
        eff.garcons = fields['garcons']
        eff.nouveaux = fields['nouveaux']
        eff.total_enfants = eff.filles + eff.garcons
        eff.total = eff.hommes + eff.femmes + eff.total_enfants + eff.nouveaux

        # Respect de l'unicité (eglise, date)
        exists = CulteEffectif.objects.exclude(id=eff.id).filter(eglise=eff.eglise, date=eff.date).first()
        if exists:
            return JsonResponse({'success': False, 'error': "Un enregistrement existe déjà pour cette église et cette date."}, status=409)

        eff.save()
        return JsonResponse({'success': True, 'id': eff.id, 'total': eff.total, 'date': eff.date.strftime('%Y-%m-%d')})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

def api_membres(request):
    """API: retourne la liste des membres au format JSON."""
    try:
        qs = Membre.objects.select_related('eglise').all().order_by('nom', 'prenom')

        def fsize(f):
            try:
                return f.size if f else None
            except Exception:
                return None

        data = []
        for m in qs:
            data.append({
                'id': m.id,
                'nom': m.nom,
                'prenom': m.prenom,
                'sexe': m.sexe,
                'date_naissance': m.date_naissance.isoformat() if m.date_naissance else None,
                'lieu_naissance': m.lieu_naissance or '',
                'adresse': m.adresse or '',
                'telephone': m.telephone or '',
                'email': m.email or '',
                'profession': m.profession or '',
                'etat_civil': m.etat_civil or '',
                'nombre_enfants': m.nombre_enfants or 0,
                'date_bapteme': m.date_bapteme.isoformat() if m.date_bapteme else None,
                'lieu_bapteme': m.lieu_bapteme or '',
                'date_adhesion': m.date_adhesion.isoformat() if m.date_adhesion else None,
                'motivation': m.motivation or '',
                'services': m.services or '',
                'departement': getattr(m, 'departement', '') or '',
                'soutien_financier': bool(getattr(m, 'soutien_financier', False)),
                'montant_souhaite': str(getattr(m, 'montant_souhaite', '') or '') or None,
                'certificat_bapteme_url': m.certificat_bapteme.url if getattr(m, 'certificat_bapteme', None) else '',
                'certificat_bapteme_size': fsize(getattr(m, 'certificat_bapteme', None)),
                'eglise': {
                    'id': m.eglise.id if m.eglise else None,
                    'nom': m.eglise.nom if m.eglise else None,
                }
            })
        return JsonResponse({'success': True, 'count': len(data), 'membres': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def modifier_membre(request, membre_id: int):
    """Edition d'un membre avec ModelForm (tous les champs)."""
    membre = get_object_or_404(Membre, id=membre_id)
    if request.method == 'POST':
        form = MembreForm(request.POST, request.FILES, instance=membre)
        if form.is_valid():
            form.save()
            messages.success(request, "Modification enregistrée.")
            return redirect('main:details_membre', membre_id=membre.id)
        else:
            messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = MembreForm(instance=membre)
    return render(request, 'main/membre_modifier.html', {'membre': membre, 'form': form})

@login_required
@ensure_csrf_cookie
def departements(request):
    """Vue pour la page Nos Départements: formulaire + tableau dynamique."""
    eglises = Eglise.objects.all().order_by('nom')
    if request.method == 'POST':
        try:
            # Mode 1: si current_dept_id est fourni => on rattache simplement les fichiers au département existant
            current_id_raw = (request.POST.get('current_dept_id') or '').strip()
            if current_id_raw.isdigit():
                dep = get_object_or_404(Departement, id=int(current_id_raw))
                # DEBUG: Afficher l'ID du département et son nom
                print(f"DEBUG[DEPT-M1] current_dept_id={current_id_raw} -> dep.id={dep.id}, dep.nom={dep.nom}")
                # En option: effacer les PJ existantes si demandé
                try:
                    clear_flag = (request.POST.get('attachments_clear') or '').strip().lower()
                    if clear_flag in ('1', 'true', 'on', 'oui', 'yes'):
                        print(f"DEBUG[DEPT-M1] attachments_clear demandé -> suppression des PJ existantes (count={dep.attachments.count()})")
                        for att in list(dep.attachments.all()):
                            try:
                                if getattr(att, 'fichier', None):
                                    att.fichier.delete(save=False)
                            except Exception:
                                pass
                            print(f"DEBUG[DEPT-M1] Suppression PJ: att.id={att.id}, nom={att.nom_original}, fichier={getattr(att, 'fichier', None)}")
                            att.delete()
                except Exception:
                    pass
                # Sauver les nouveaux fichiers
                try:
                    files = request.FILES.getlist('attachments') or request.FILES.getlist('attachments[]')
                    if not files:
                        single = request.FILES.get('document_fichier')
                        if single:
                            files = [single]
                    print(f"DEBUG[DEPT-M1] {len(files)} fichier(s) reçu(s) pour PJ (attachments)")
                    for i, f in enumerate(files):
                        print(f"DEBUG[DEPT-M1] Fichier {i+1}: name={getattr(f, 'name', '')}, size={getattr(f, 'size', '?')} bytes")
                        att = DepartementAttachment.objects.create(
                            departement=dep,
                            fichier=f,
                            nom_original=getattr(f, 'name', '') or ''
                        )
                        print(f"DEBUG[DEPT-M1] PJ créée: att.id={att.id}, nom={att.nom_original}, fichier={att.fichier.name}")
                except Exception as e:
                    print(f"DEBUG[DEPT-M1] ERREUR lors de la sauvegarde des PJ: {e}")
                messages.success(request, "Plan d’action enregistré pour le département existant.")
                return redirect('main:departements')

            # Mode 2: création d'un nouveau département (comportement existant)
            eglise_id = request.POST.get('eglise_id')
            nom = (request.POST.get('nom') or '').strip()
            if not eglise_id or not nom:
                messages.error(request, "Veuillez sélectionner l'Église et saisir le nom du département.")
                return render(request, 'main/nos_departements.html', {
                    'eglises': eglises,
                    'departements': Departement.objects.select_related('eglise').order_by('nom'),
                    'form_data': request.POST,
                })

            eglise = get_object_or_404(Eglise, id=int(eglise_id))

            # Création avec conversion minimale des champs
            def to_bool(v):
                return str(v).lower() in ('1', 'true', 'on', 'oui', 'yes')

            from datetime import datetime as _dt
            def parse_date(v):
                if not v:
                    return None
                try:
                    return _dt.strptime(v, '%Y-%m-%d').date()
                except Exception:
                    return None
            def parse_int(v, d=None):
                try:
                    return int(v)
                except Exception:
                    return d
            def parse_decimal(v):
                try:
                    from decimal import Decimal
                    return Decimal(v)
                except Exception:
                    return None

            dep = Departement.objects.create(
                eglise=eglise,
                nom=nom,
                responsable=request.POST.get('responsable') or None,
                telephone=request.POST.get('telephone') or None,
                email=request.POST.get('email') or None,
                description=request.POST.get('description') or None,
                date_creation=parse_date(request.POST.get('date_creation')),
                statut_actif=to_bool(request.POST.get('statut_actif')),
                vision=request.POST.get('vision') or None,
                mission=request.POST.get('mission') or None,
                objectifs_specifiques=request.POST.get('objectifs_specifiques') or None,
                membres_clefs=request.POST.get('membres_clefs') or None,
                nombre_membres=parse_int(request.POST.get('nombre_membres'), 0),
                ressources_disponibles=request.POST.get('ressources_disponibles') or None,
                besoins_prioritaires=request.POST.get('besoins_prioritaires') or None,
                budget_annuel_estime=parse_decimal(request.POST.get('budget_annuel_estime')),
                activites_majeures=request.POST.get('activites_majeures') or None,
                prochaine_activite_date=parse_date(request.POST.get('prochaine_activite_date')),
                prochaine_activite_lieu=request.POST.get('prochaine_activite_lieu') or None,
                valide_par_conseil=to_bool(request.POST.get('valide_par_conseil')),
                date_validation=parse_date(request.POST.get('date_validation')),
                observations_conseil=request.POST.get('observations_conseil') or None,
            )

            # Sauvegarder les pièces jointes si fournies (création)
            try:
                files = request.FILES.getlist('attachments') or request.FILES.getlist('attachments[]')
                # Gérer aussi le champ fichier unique provenant de l'input "document_fichier"
                if not files:
                    single = request.FILES.get('document_fichier')
                    if single:
                        files = [single]
                
                # DEBUG: Afficher les fichiers reçus
                print(f"DEBUG: Fichiers reçus pour département {dep.nom}: {len(files)} fichier(s)")
                for i, f in enumerate(files):
                    print(f"DEBUG: Fichier {i+1}: {f.name} ({f.size} bytes)")
                
                for f in files:
                    attachment = DepartementAttachment.objects.create(
                        departement=dep,
                        fichier=f,
                        nom_original=getattr(f, 'name', '') or ''
                    )
                    print(f"DEBUG: Attachment créé: ID={attachment.id}, fichier={attachment.fichier.name}")
                    
            except Exception as e:
                # DEBUG: Afficher l'erreur
                print(f"DEBUG: Erreur lors de la sauvegarde des fichiers: {e}")
                # En cas de problème de fichier, ignorer silencieusement pour ne pas bloquer l'enregistrement
                pass

            messages.success(request, "Département enregistré avec succès.")
            return redirect('main:departements')
        except Exception as e:
            # Ne pas afficher d'erreur technique détaillée à l'utilisateur
            messages.error(request, "Une erreur est survenue lors de l'enregistrement du département.")

    departements_qs = Departement.objects.select_related('eglise').prefetch_related('attachments').order_by('nom')
    return render(request, 'main/nos_departements.html', {
        'eglises': eglises,
        'departements': departements_qs,
    })

def voir_departements(request):
    """Vue pour afficher tous les départements dans une page dédiée."""
    departements_qs = Departement.objects.select_related('eglise').prefetch_related('attachments').order_by('nom')
    return render(request, 'main/voir_departements.html', {
        'departements': departements_qs,
    })


@require_POST
def supprimer_departement_attachment(request, attachment_id: int):
    """Supprime une pièce jointe de département (fichier + enregistrement).

    - Requiert POST (CSRF protégé par défaut)
    - Réponse JSON si AJAX (header X-Requested-With)
    - Sinon, messages + redirection vers la page des départements
    """
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    try:
        att = get_object_or_404(DepartementAttachment, id=attachment_id)
        dep_id = att.departement_id

        # Supprimer le fichier du stockage si présent
        try:
            if getattr(att, 'fichier', None):
                att.fichier.delete(save=False)
        except Exception:
            # Ne pas bloquer la suppression en cas d'erreur de fichier
            pass

        att.delete()
        if is_ajax:
            return JsonResponse({'success': True, 'deleted_id': attachment_id})
        messages.success(request, "Pièce jointe supprimée avec succès.")
    except Exception as e:
        if is_ajax:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        messages.error(request, f"Erreur lors de la suppression de la pièce jointe: {str(e)}")

    return redirect('main:departements')

def modifier_departement(request, departement_id: int):
    """Met à jour un département existant.

    - Accepte uniquement POST (AJAX recommandé) avec des champs mappés depuis la modale.
    - Retourne JSON { success: True, departement: { ...champs utiles pour mise à jour du tableau... } }
    """
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    if request.method != 'POST':
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'}, status=405)
        messages.error(request, "Méthode non autorisée pour la modification du département.")
        return redirect('main:departements')

    try:
        dep = get_object_or_404(Departement, id=departement_id)

        # Aides de parsing
        from decimal import Decimal
        def parse_decimal(v):
            try:
                if v is None or v == '':
                    return None
                # Nettoyer '1 000 GNF' ou '1000' -> Decimal('1000')
                cleaned = ''.join(ch for ch in str(v) if ch.isdigit() or ch in ',.')
                cleaned = cleaned.replace(',', '.')
                return Decimal(cleaned)
            except Exception:
                return None
        from datetime import datetime as _dt
        def parse_date(v):
            if not v:
                return None
            for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
                try:
                    return _dt.strptime(v, fmt).date()
                except Exception:
                    continue
            return None

        # Mappage des champs reçus du formulaire de la modale
        # Champs attendus côté JS (si présents):
        # nom, responsable, telephone, email, description, mission, activites_regulieres,
        # membres_comite, budget_annuel_estime, materiel, besoins, autres_besoins,
        # frequence, evenements, approuve_par, date_validation
        dep.nom = request.POST.get('nom', dep.nom)
        dep.responsable = request.POST.get('responsable', dep.responsable)
        dep.telephone = request.POST.get('telephone', dep.telephone)
        dep.email = request.POST.get('email', dep.email)
        dep.description = request.POST.get('description', dep.description)
        dep.mission = request.POST.get('mission', dep.mission)

        activites_regulieres = request.POST.get('activites_regulieres')
        if activites_regulieres:
            # Stocker dans activites_majeures si fourni
            dep.activites_majeures = activites_regulieres

        membres_comite = request.POST.get('membres_comite')
        if membres_comite:
            dep.membres_clefs = membres_comite

        materiel = request.POST.get('materiel')
        if materiel:
            dep.ressources_disponibles = materiel

        besoins = request.POST.get('besoins')
        autres_besoins = request.POST.get('autres_besoins')
        if besoins or autres_besoins:
            merged = ', '.join([p for p in [besoins, autres_besoins] if p])
            dep.besoins_prioritaires = merged or dep.besoins_prioritaires

        frequence = request.POST.get('frequence')
        evenements = request.POST.get('evenements')
        if frequence or evenements:
            # Concat simple pour consigner le calendrier dans observations si pas de modèle dédié
            cal = []
            if frequence:
                cal.append(f"Fréquence: {frequence}")
            if evenements:
                cal.append(f"Événements: {evenements}")
            joined = ' | '.join(cal)
            dep.observations_conseil = (joined if not dep.observations_conseil else f"{dep.observations_conseil} | {joined}")

        budget_raw = request.POST.get('budget_annuel_estime') or request.POST.get('budget')
        if budget_raw is not None:
            dep.budget_annuel_estime = parse_decimal(budget_raw)

        dep.date_validation = parse_date(request.POST.get('date_validation')) or dep.date_validation

        # Optionnel: si 'approuve_par' est fourni, on l'ajoute aux observations
        approuve_par = request.POST.get('approuve_par')
        if approuve_par:
            info = f"Approuvé par: {approuve_par}"
            dep.observations_conseil = (info if not dep.observations_conseil else f"{dep.observations_conseil} | {info}")

        dep.save()

        # Si demandé: effacer toutes les pièces jointes existantes
        try:
            clear_flag = (request.POST.get('attachments_clear') or '').strip().lower()
            if clear_flag in ('1', 'true', 'on', 'oui', 'yes'):
                for att in list(dep.attachments.all()):
                    try:
                        if getattr(att, 'fichier', None):
                            att.fichier.delete(save=False)
                    except Exception:
                        # Ignorer les erreurs liées au stockage
                        pass
                    att.delete()
        except Exception:
            # Ne pas bloquer la suite en cas d'erreur lors du nettoyage
            pass

        # Ajouter de nouvelles pièces jointes si envoyées
        try:
            files = request.FILES.getlist('attachments') or request.FILES.getlist('attachments[]')
            # Gérer aussi le champ fichier unique provenant de l'input "document_fichier"
            if not files:
                single = request.FILES.get('document_fichier')
                if single:
                    files = [single]
            for f in files:
                DepartementAttachment.objects.create(
                    departement=dep,
                    fichier=f,
                    nom_original=getattr(f, 'name', '') or ''
                )
        except Exception:
            pass

        # Réponse avec données formatées pour la mise à jour du tableau
        def fmt_budget(val):
            try:
                return f"{int(val):,} GNF".replace(',', ' ')
            except Exception:
                try:
                    return f"{val} GNF" if val is not None else ''
                except Exception:
                    return ''
        def fmt_date(d):
            return d.strftime('%d/%m/%Y') if d else ''

        data = {
            'id': dep.id,
            'nom': dep.nom or '',
            'responsable': dep.responsable or '',
            'telephone': dep.telephone or '',
            'email': dep.email or '',
            'description': dep.description or '',
            'mission': dep.mission or '',
            'membres_comite': dep.membres_clefs or '',
            'budget': fmt_budget(dep.budget_annuel_estime),
            'materiel': dep.ressources_disponibles or '',
            'besoins': dep.besoins_prioritaires or '',
            'frequence': frequence or '',
            'evenements': evenements or '',
            'approuve_par': approuve_par or '',
            'date_validation': fmt_date(dep.date_validation),
        }

        if is_ajax:
            return JsonResponse({'success': True, 'departement': data})
        messages.success(request, "Département modifié avec succès.")
        return redirect('main:departements')
    except Exception as e:
        if is_ajax:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        messages.error(request, "Erreur lors de la modification du département.")
        return redirect('main:departements')

@login_required
@ensure_csrf_cookie
def finances(request):
    """Vue pour la page Finances

    - GET: affiche la page avec la liste des églises
    - POST: enregistre les lignes de recettes et de dépenses (modèle `Finance`)
            et sauvegarde les pièces jointes dans le stockage.
    """
    eglises = Eglise.objects.all().order_by('nom')
    # Préparer la liste des opérations pour affichage dans le tableau
    operations = (
        Finance.objects
        .select_related('eglise', 'report')
        .prefetch_related('report__attachments')
        .all()
        .order_by('-date_operation', '-id')
    )
    reports = FinanceReport.objects.select_related('eglise').prefetch_related('attachments').all().order_by('-date_rapport', '-id')

    if request.method != 'POST':
        return render(request, 'main/finances.html', {
            'eglises': eglises,
            'operations': operations,
            'reports': reports,
        })

    # Helpers de parsing
    def parse_decimal(val):
        if val is None:
            return None
        s = str(val).replace('GNF', '').replace('\xa0', ' ').replace(',', ' ').strip()
        s = s.replace(' ', '')
        if not s:
            return None
        try:
            # limiter à 2 décimales
            return Decimal(s)
        except (InvalidOperation, ValueError):
            return None

    def parse_date(val):
        if not val:
            return None
        try:
            return datetime.strptime(val, '%Y-%m-%d').date()
        except Exception:
            return None

    def clean_str(val):
        return (val or '').strip()

    try:
        # 1) Résoudre l'église par identifiant (champ `eglise_id`)
        eglise_id = clean_str(request.POST.get('eglise_id'))
        if not eglise_id:
            messages.error(request, "Veuillez sélectionner l'Église.")
            return render(request, 'main/finances.html', { 'eglises': eglises, 'operations': operations })
        try:
            eglise = Eglise.objects.get(id=eglise_id)
        except Eglise.DoesNotExist:
            messages.error(request, "Église introuvable. Veuillez réessayer.")
            return render(request, 'main/finances.html', { 'eglises': eglises, 'operations': operations })

        # 2) Récupérer les champs d'en-tête (rapport)
        date_rapport = clean_str(request.POST.get('date_rapport')) or None
        periode_du = clean_str(request.POST.get('periode_du')) or None
        periode_au = clean_str(request.POST.get('periode_au')) or None
        responsable = clean_str(request.POST.get('responsable_financier') or request.POST.get('responsable'))
        contact = clean_str(request.POST.get('contact_financier') or request.POST.get('contact'))
        email = clean_str(request.POST.get('email_financier') or request.POST.get('email'))
        verifie_par = clean_str(request.POST.get('verifie_par'))
        date_verification = clean_str(request.POST.get('date_verification')) or None
        approuve_par = clean_str(request.POST.get('approuve_par'))

        # Créer le rapport maintenant (totaux remplis après création des lignes)
        report = FinanceReport.objects.create(
            eglise=eglise,
            date_rapport=parse_date(date_rapport),
            periode_du=parse_date(periode_du),
            periode_au=parse_date(periode_au),
            responsable=responsable,
            contact=contact,
            email=email,
            verifie_par=verifie_par,
            date_verification=parse_date(date_verification),
            approuve_par=approuve_par,
        )

        # 3) Récupérer les listes dynamiques
        r_sources = request.POST.getlist('revenu_source[]')
        r_montants = request.POST.getlist('revenu_montant[]')
        r_dates = request.POST.getlist('revenu_date[]')
        r_modes = request.POST.getlist('revenu_mode[]')
        r_notes = request.POST.getlist('revenu_remarques[]')

        d_natures = request.POST.getlist('depense_nature[]')
        d_montants = request.POST.getlist('depense_montant[]')
        d_benef = request.POST.getlist('depense_beneficiaire[]')
        d_just = request.POST.getlist('depense_justificatif[]')
        d_categ = request.POST.getlist('depense_categorie[]')

        created_count = 0
        total_recettes = Decimal('0')
        total_depenses = Decimal('0')

        # 4) Enregistrer les recettes
        max_r = max(len(r_sources), len(r_montants), len(r_dates), len(r_modes), len(r_notes)) if any([r_sources, r_montants, r_dates, r_modes, r_notes]) else 0
        skipped_recettes = 0
        for i in range(max_r):
            libelle = clean_str(r_sources[i]) if i < len(r_sources) else ''
            montant = parse_decimal(r_montants[i]) if i < len(r_montants) else None
            dte = parse_date(r_dates[i]) if i < len(r_dates) else None
            mode = clean_str(r_modes[i]) if i < len(r_modes) else ''
            note = clean_str(r_notes[i]) if i < len(r_notes) else ''

            # Filtrer les lignes vides
            if not libelle and (montant is None or montant == 0):
                continue
            if montant is None or montant <= 0:
                messages.error(request, f"Montant invalide (doit être > 0) pour une ligne de recette (ligne {i+1}).")
                return render(request, 'main/finances.html', { 'eglises': eglises, 'operations': operations })

            # Anti-doublon: même église, même date, même montant, même type, même libellé (insensible à la casse)
            try:
                if Finance.objects.filter(
                    eglise=eglise,
                    type_operation=Finance.TYPE_RECETTE,
                    montant=montant,
                    date_operation=dte or datetime.today().date(),
                    libelle__iexact=libelle or 'Recette',
                ).exists():
                    skipped_recettes += 1
                    continue
            except Exception:
                # En cas d'erreur de vérification, on continue la création pour ne pas bloquer
                pass

            Finance.objects.create(
                report=report,
                eglise=eglise,
                type_operation=Finance.TYPE_RECETTE,
                montant=montant,
                date_operation=dte or datetime.today().date(),
                libelle=libelle or 'Recette',
                description=note or (f"Mode: {mode}" if mode else ''),
                beneficiaire='',
                categorie=mode or 'Recette',
                source_revenu=libelle,
                mode_paiement=mode,
                remarques=note,
            )
            created_count += 1
            try:
                total_recettes += montant or Decimal('0')
            except Exception:
                pass

        # 5) Enregistrer les dépenses
        max_d = max(len(d_natures), len(d_montants), len(d_benef), len(d_just), len(d_categ)) if any([d_natures, d_montants, d_benef, d_just, d_categ]) else 0
        skipped_depenses = 0
        for i in range(max_d):
            nature = clean_str(d_natures[i]) if i < len(d_natures) else ''
            montant = parse_decimal(d_montants[i]) if i < len(d_montants) else None
            benef = clean_str(d_benef[i]) if i < len(d_benef) else ''
            justificatif = clean_str(d_just[i]) if i < len(d_just) else ''
            categorie = clean_str(d_categ[i]) if i < len(d_categ) else ''

            if not nature and (montant is None or montant == 0):
                continue
            if montant is None or montant <= 0:
                messages.error(request, f"Montant invalide (doit être > 0) pour une ligne de dépense (ligne {i+1}).")
                return render(request, 'main/finances.html', { 'eglises': eglises, 'operations': operations })

            # Anti-doublon: même église, même date (aujourd'hui), même montant, même type, même libellé (nature)
            try:
                if Finance.objects.filter(
                    eglise=eglise,
                    type_operation=Finance.TYPE_DEPENSE,
                    montant=montant,
                    date_operation=datetime.today().date(),
                    libelle__iexact=nature or 'Dépense',
                ).exists():
                    skipped_depenses += 1
                    continue
            except Exception:
                pass

            Finance.objects.create(
                report=report,
                eglise=eglise,
                type_operation=Finance.TYPE_DEPENSE,
                montant=montant,
                date_operation=datetime.today().date(),
                libelle=nature or 'Dépense',
                description=justificatif,
                beneficiaire=benef,
                categorie=categorie or 'Dépense',
                nature_depense=nature,
                justificatif_numero=justificatif,
            )
            created_count += 1
            try:
                total_depenses += montant or Decimal('0')
            except Exception:
                pass

        # 6) Sauvegarder les pièces jointes (modèle FinanceAttachment)
        try:
            files = request.FILES.getlist('pieces_jointes')
            if files:
                for f in files:
                    try:
                        FinanceAttachment.objects.create(report=report, fichier=f)
                    except Exception:
                        continue
        except Exception:
            # Ne pas bloquer l'enregistrement si la sauvegarde de fichiers échoue
            pass

        # 7) Calculer les totaux et le solde
        try:
            solde_initial = parse_decimal(request.POST.get('solde_initial')) or Decimal('0')
        except Exception:
            solde_initial = Decimal('0')
        solde_final = solde_initial + total_recettes - total_depenses

        try:
            report.total_recettes = total_recettes
            report.total_depenses = total_depenses
            report.solde_initial = solde_initial
            report.solde_final = solde_final
            try:
                report.nombre_pieces_jointes = report.attachments.count()
            except Exception:
                pass
            report.save(update_fields=['total_recettes','total_depenses','solde_initial','solde_final','nombre_pieces_jointes'])
        except Exception:
            pass

        if created_count == 0:
            messages.warning(request, "Aucune ligne de recette ou de dépense valide n'a été détectée.")
            return render(request, 'main/finances.html', { 'eglises': eglises, 'operations': operations })

        # Informer si certaines lignes ont été ignorées pour cause de doublon
        if skipped_recettes or skipped_depenses:
            messages.warning(request, f"{skipped_recettes} ligne(s) de recette et {skipped_depenses} ligne(s) de dépense ont été ignorées car déjà présentes (doublons).")
        messages.success(request, "Les données financières ont été enregistrées avec succès.")
        return redirect('main:finances')
    except Exception as e:
        if getattr(settings, 'DEBUG', False):
            messages.error(request, f"Une erreur est survenue lors de l'enregistrement des finances: {e}")
        else:
            messages.error(request, "Une erreur est survenue lors de l'enregistrement des finances.")
        return render(request, 'main/finances.html', { 'eglises': eglises, 'operations': operations })

def voir_personnel_pastoral(request):
    """Affiche le tableau complet du personnel pastoral avec outils d'export."""
    pasteurs = PersonnelPastoral.objects.all().order_by('-id')
    eglises = Eglise.objects.all().order_by('nom')
    return render(request, 'main/table_personnel_pastoral.html', {
        'pasteurs': pasteurs,
        'eglises': eglises,
    })

def actualite(request):
    """Vue pour la page Actualité"""
    return render(request, 'main/actualite.html')

def evenements(request):
    """Vue pour la page Événements"""
    eglises = Eglise.objects.all().order_by('nom')
    return render(request, 'main/evenements.html', {
        'eglises': eglises,
    })

@login_required
def effectifs_culte_complet(request):
    """Page complète listant tous les effectifs du culte avec filtres et exports."""
    eglises = Eglise.objects.all().order_by('nom')
    return render(request, 'main/effectifs_culte_complet.html', {
        'eglises': eglises,
    })

@ensure_csrf_cookie
def planification(request):
    """Vue pour la page Planification"""
    return render(request, 'main/planification.html')

def suivi_presences(request):
    """Vue pour la page Suivi des présences"""
    return render(request, 'main/suivi_presences.html')

def archivage_rapports(request):
    """Vue pour la page Archivage des rapports d'activités"""
    return render(request, 'main/archivage_rapports.html')

def communication(request):
    """Vue pour la page Communication"""
    return render(request, 'main/communication.html')

def biographie_pasteurs(request):
    """Vue pour la page Biographie des Pasteurs"""
    return render(request, 'main/biographie_pasteurs.html')

def statut_reglements(request):
    """Vue pour afficher le PDF Statut et Règlements de manière sécurisée"""
    context = {
        'pdf_title': 'Statuts et Règlements',
        'pdf_path': 'main/images/STATUT_REGLEMENTS.pdf'
    }
    return render(request, 'main/pdf_viewer.html', context)

def plan_strategique(request):
    """Vue pour afficher le PDF Plan Stratégique de manière sécurisée"""
    context = {
        'pdf_title': 'Plan Stratégique',
        'pdf_path': 'main/images/Plan_strategique.pdf'
    }
    return render(request, 'main/pdf_viewer.html', context)

def contact(request):
    """Vue pour la page Nous contacter (GET) et traitement du formulaire (POST)."""
    if request.method == 'POST':
        name = (request.POST.get('name') or '').strip()
        email = (request.POST.get('email') or '').strip()
        subject = (request.POST.get('subject') or '').strip()
        message = (request.POST.get('message') or '').strip()
        telephone = (request.POST.get('telephone') or '').strip()

        form_data = { 'name': name, 'email': email, 'subject': subject, 'message': message, 'telephone': telephone }
        errors = []

        if len(name) < 3:
            errors.append("Le nom doit contenir au moins 3 caractères.")
        if len(subject) < 3:
            errors.append("Le sujet doit contenir au moins 3 caractères.")
        if len(message) < 20:
            errors.append("Le message doit contenir au moins 20 caractères.")
        try:
            validate_email(email)
        except ValidationError:
            errors.append("Adresse email invalide.")
        # Téléphone optionnel: valider si présent (autoriser +, chiffres, espaces, tirets, parenthèses)
        if telephone:
            if not re.fullmatch(r"[+()\d\s-]{8,20}", telephone):
                errors.append("Numéro de téléphone invalide (utilisez chiffres, +, espaces, - ou ()).")

        if errors:
            for e in errors:
                messages.error(request, e)
            # PRG: stocker les données en session puis rediriger vers GET
            try:
                request.session['contact_form_data'] = form_data
            except Exception:
                pass
            return redirect('main:contact')

        # Envoi d'email
        try:
            subject_line = f"[Contact Église] {subject}"
            body = (
                "Nouveau message de contact\n\n"
                f"Nom: {name}\n"
                f"Email: {email}\n"
                f"Téléphone: {telephone or '—'}\n"
                f"Sujet: {subject}\n\n"
                f"Message:\n{message}\n"
            )
            send_mail(
                subject_line,
                body,
                settings.DEFAULT_FROM_EMAIL,
                [getattr(settings, 'CONTACT_RECEIVER', settings.DEFAULT_FROM_EMAIL)],
                fail_silently=False,
            )
            messages.success(request, "Merci, votre message a été envoyé. Nous vous répondrons sous 24 à 48 heures.")
            return redirect('main:contact')
        except Exception as e:
            messages.error(request, "Une erreur est survenue lors de l'envoi de l'email. Veuillez réessayer plus tard.")
            try:
                request.session['contact_form_data'] = form_data
            except Exception:
                pass
            return redirect('main:contact')
    # GET: restaurer d'éventuelles données depuis la session (PRG)
    form_data = {}
    try:
        stored = request.session.pop('contact_form_data', None)
        if stored:
            form_data = stored
    except Exception:
        pass
    return render(request, 'main/contact.html', { 'form_data': form_data })

def get_eglise_by_name(request):
    """Vue pour récupérer les détails d'une église par son nom"""
    if request.method == 'GET':
        nom_eglise = request.GET.get('nom', '')
        
        if nom_eglise:
            try:
                eglise = Eglise.objects.get(nom__icontains=nom_eglise)
                data = {
                    'id': eglise.id,
                    'nom': eglise.nom,
                    'ville': eglise.ville,
                    'pays': eglise.pays,
                    'responsable': eglise.responsable,
                    'telephone': eglise.telephone,
                    'email': eglise.email,
                }
                return JsonResponse(data)
            except Eglise.DoesNotExist:
                return JsonResponse({'error': 'Église non trouvée'}, status=404)
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

# ============================
# API Activités hebdomadaires
# ============================
from django.views.decorators.http import require_http_methods
from django.forms.models import model_to_dict
from .models import ActiviteHebdo

def _parse_json(request):
    try:
        body = request.body.decode('utf-8') if request.body else ''
        return json.loads(body) if body else {}
    except Exception:
        return {}

def _serialize_hebdo(a: ActiviteHebdo):
    d = model_to_dict(a, fields=['id','nom','jour','heure_debut','duree_minutes','heure_fin','lieu','responsable','description'])
    d['heure_debut'] = a.heure_debut.strftime('%H:%M') if a.heure_debut else ''
    d['heure_fin'] = a.heure_fin.strftime('%H:%M') if a.heure_fin else ''
    d['eglise_id'] = getattr(getattr(a, 'eglise', None), 'id', None)
    d['eglise_nom'] = getattr(getattr(a, 'eglise', None), 'nom', '')
    return d

@require_http_methods(["GET"])
def api_hebdo_list(request):
    qs = ActiviteHebdo.objects.all().order_by('jour','heure_debut','nom')
    eg = request.GET.get('eglise_id')
    if eg:
        try:
            qs = qs.filter(eglise_id=int(eg))
        except Exception:
            pass
    q = (request.GET.get('q') or '').strip()
    if q:
        from django.db.models import Q
        qs = qs.filter(Q(nom__icontains=q) | Q(lieu__icontains=q) | Q(responsable__icontains=q) | Q(description__icontains=q))
    items = [_serialize_hebdo(x) for x in qs]
    return JsonResponse({'success': True, 'items': items})

@require_http_methods(["POST"]) 
def api_hebdo_create(request):
    data = _parse_json(request)
    try:
        nom = (data.get('nom') or '').strip()
        jour = (data.get('jour') or '').strip()
        heure_debut = (data.get('heure_debut') or '').strip()
        duree = int(data.get('duree_minutes') or 0)
        lieu = (data.get('lieu') or '').strip()
        responsable = (data.get('responsable') or '').strip()
        description = (data.get('description') or '').strip()
        eglise_id = data.get('eglise_id')
        if not nom or not jour or not heure_debut:
            return JsonResponse({'success': False, 'error': 'Champs requis manquants.'}, status=400)
        from datetime import datetime as _dt
        hd = _dt.strptime(heure_debut, '%H:%M').time()
        hf = None
        if duree and duree > 0:
            # calcul heure fin
            total = hd.hour*60 + hd.minute + duree
            total = total % (24*60)
            from datetime import time as _time
            hf = _time(total//60, total%60)
        eg_obj = None
        if eglise_id:
            try:
                eg_obj = Eglise.objects.get(id=int(eglise_id))
            except Exception:
                eg_obj = None
        obj = ActiviteHebdo.objects.create(
            nom=nom, jour=jour, heure_debut=hd, duree_minutes=max(0,duree), heure_fin=hf,
            lieu=lieu, responsable=responsable, description=description, eglise=eg_obj
        )
        return JsonResponse({'success': True, 'item': _serialize_hebdo(obj)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@require_http_methods(["POST"]) 
def api_hebdo_update(request, item_id):
    data = _parse_json(request)
    try:
        obj = ActiviteHebdo.objects.get(id=item_id)
        nom = data.get('nom')
        jour = data.get('jour')
        heure_debut = data.get('heure_debut')
        duree = data.get('duree_minutes')
        lieu = data.get('lieu')
        responsable = data.get('responsable')
        description = data.get('description')
        eglise_id = data.get('eglise_id')
        from datetime import datetime as _dt, time as _time
        if nom is not None: obj.nom = (nom or '').strip()
        if jour is not None: obj.jour = (jour or '').strip()
        if heure_debut is not None and (heure_debut or '').strip():
            obj.heure_debut = _dt.strptime(heure_debut, '%H:%M').time()
        if duree is not None:
            try:
                obj.duree_minutes = max(0, int(duree))
            except Exception:
                obj.duree_minutes = 0
        if obj.heure_debut and obj.duree_minutes:
            total = obj.heure_debut.hour*60 + obj.heure_debut.minute + obj.duree_minutes
            total = total % (24*60)
            obj.heure_fin = _time(total//60, total%60)
        if lieu is not None: obj.lieu = (lieu or '').strip()
        if responsable is not None: obj.responsable = (responsable or '').strip()
        if description is not None: obj.description = (description or '').strip()
        if eglise_id is not None:
            try:
                obj.eglise = Eglise.objects.get(id=int(eglise_id))
            except Exception:
                obj.eglise = None
        obj.save()
        return JsonResponse({'success': True, 'item': _serialize_hebdo(obj)})
    except ActiviteHebdo.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Activité introuvable.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@require_http_methods(["POST"]) 
def api_hebdo_delete(request, item_id):
    try:
        obj = ActiviteHebdo.objects.get(id=item_id)
        obj.delete()
        return JsonResponse({'success': True})
    except ActiviteHebdo.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Activité introuvable.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

@require_http_methods(["GET"])
def api_eglises_list(request):
    data = list(Eglise.objects.all().order_by('nom').values('id','nom'))
    return JsonResponse({'success': True, 'items': data})

@require_http_methods(["GET"])
def api_departements_list(request):
    try:
        qs = Departement.objects.select_related('eglise').filter(statut_actif=True).order_by('nom')
        eg = request.GET.get('eglise_id')
        if eg:
            try:
                qs = qs.filter(eglise_id=int(eg))
            except Exception:
                pass
        items = [{'id': d.id, 'nom': d.nom, 'eglise_id': d.eglise_id, 'eglise_nom': getattr(d.eglise,'nom','') } for d in qs]
        return JsonResponse({'success': True, 'items': items})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

# Exports activités hebdo
@require_http_methods(["GET"])
def api_hebdo_export_csv(request):
    try:
        qs = ActiviteHebdo.objects.all().order_by('jour','heure_debut','nom')
        eg = request.GET.get('eglise_id')
        if eg:
            try:
                qs = qs.filter(eglise_id=int(eg))
            except Exception:
                pass
        q = (request.GET.get('q') or '').strip()
        if q:
            from django.db.models import Q
            qs = qs.filter(Q(nom__icontains=q) | Q(lieu__icontains=q) | Q(responsable__icontains=q) | Q(description__icontains=q))

        import csv
        from django.http import HttpResponse
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="activites_hebdo.csv"'
        response.write('\ufeff')  # BOM UTF-8 pour Excel
        w = csv.writer(response, delimiter=';')
        def clean(v):
            return _fix_mojibake(_to_str(v))
        w.writerow(['Église','Activité','Jour','Heure début','Heure fin','Lieu','Responsable','Description','Durée (min)'])
        for a in qs:
            eglise_nom = clean(getattr(getattr(a,'eglise',None),'nom',''))
            nom = clean(a.nom)
            jour = clean(a.jour)
            hd = a.heure_debut.strftime('%H:%M') if a.heure_debut else ''
            hf = a.heure_fin.strftime('%H:%M') if a.heure_fin else ''
            lieu = clean(a.lieu)
            responsable = clean(a.responsable)
            description = clean(a.description)
            duree = a.duree_minutes
            w.writerow([eglise_nom, nom, jour, hd, hf, lieu, responsable, description, duree])
        return response
    except Exception as e:
        return HttpResponse(f"Erreur export CSV: {e}", status=500)

@require_http_methods(["GET"])
def api_hebdo_export_xlsx(request):
    try:
        qs = ActiviteHebdo.objects.all().order_by('jour','heure_debut','nom')
        eg = request.GET.get('eglise_id')
        if eg:
            try:
                qs = qs.filter(eglise_id=int(eg))
            except Exception:
                pass
        q = (request.GET.get('q') or '').strip()
        if q:
            from django.db.models import Q
            qs = qs.filter(Q(nom__icontains=q) | Q(lieu__icontains=q) | Q(responsable__icontains=q) | Q(description__icontains=q))
        try:
            from openpyxl import Workbook
        except Exception:
            return api_hebdo_export_csv(request)
        wb = Workbook(); ws = wb.active; ws.title = 'Hebdo'
        headers = ['Église','Activité','Jour','Heure début','Heure fin','Lieu','Responsable','Description','Durée (min)']
        ws.append(headers)
        for a in qs:
            ws.append([
                getattr(getattr(a,'eglise',None),'nom',''),
                a.nom,
                a.jour,
                a.heure_debut.strftime('%H:%M') if a.heure_debut else '',
                a.heure_fin.strftime('%H:%M') if a.heure_fin else '',
                a.lieu,
                a.responsable,
                a.description,
                a.duree_minutes
            ])
        # largeur colonnes
        for col in ws.columns:
            max_len=0; letter=col[0].column_letter
            for c in col:
                v=str(c.value) if c.value is not None else ''
                if len(v)>max_len: max_len=len(v)
            ws.column_dimensions[letter].width=min(40,max(10,max_len+2))
        from io import BytesIO
        from django.http import HttpResponse
        bio=BytesIO(); wb.save(bio); bio.seek(0)
        resp=HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition']='attachment; filename="activites_hebdo.xlsx"'
        return resp
    except Exception as e:
        from django.http import HttpResponse
        return HttpResponse(f"Erreur export XLSX: {e}", status=500)

# ===================== Planification Stratégique (StrategicPlan) =====================
def _serialize_sp(obj: StrategicPlan):
    return {
        'id': obj.id,
        'eglise_id': getattr(getattr(obj, 'eglise', None), 'id', None),
        'eglise_nom': getattr(getattr(obj, 'eglise', None), 'nom', ''),
        'departement': obj.departement or '',
        'type': obj.type_planification,
        'activite': obj.activite,
        'jour': obj.jour,
        'date_prev': obj.date_prev.strftime('%Y-%m-%d') if obj.date_prev else '',
        'heure': obj.heure.strftime('%H:%M') if obj.heure else '',
        'lieu': obj.lieu or '',
        'besoins': obj.besoins or '',
        'responsable': obj.responsable or '',
        'objectif': obj.objectif or '',
        'budget': str(obj.budget) if obj.budget is not None else '',
        'statut': obj.statut,
        'nom_resp': obj.nom_resp,
        'nom_pasteur': obj.nom_pasteur,
        'date_valid': obj.date_valid.strftime('%Y-%m-%d') if obj.date_valid else '',
        'created_at': obj.created_at.strftime('%Y-%m-%d %H:%M') if obj.created_at else ''
    }

@require_http_methods(["GET"])
def api_sp_list(request):
    try:
        qs = StrategicPlan.objects.select_related('eglise').all().order_by('-date_prev', 'heure', 'activite')
        eg = request.GET.get('eglise_id')
        if eg:
            try:
                qs = qs.filter(eglise_id=int(eg))
            except Exception:
                pass
        # Filtres de période (date_prev)
        df = (request.GET.get('date_from') or '').strip()
        dt = (request.GET.get('date_to') or '').strip()
        if df:
            try:
                from datetime import datetime as _dt
                qs = qs.filter(date_prev__gte=_dt.strptime(df, '%Y-%m-%d').date())
            except Exception:
                pass
        if dt:
            try:
                from datetime import datetime as _dt
                qs = qs.filter(date_prev__lte=_dt.strptime(dt, '%Y-%m-%d').date())
            except Exception:
                pass
        q = (request.GET.get('q') or '').strip()
        if q:
            from django.db.models import Q
            qs = qs.filter(
                Q(activite__icontains=q) | Q(departement__icontains=q) | Q(lieu__icontains=q)
                | Q(besoins__icontains=q) | Q(responsable__icontains=q) | Q(objectif__icontains=q)
                | Q(nom_resp__icontains=q) | Q(nom_pasteur__icontains=q)
            )
        items = [_serialize_sp(x) for x in qs]
        return JsonResponse({'success': True, 'items': items})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@require_http_methods(["POST"]) 
def api_sp_create(request):
    data = _parse_json(request)
    try:
        required = ['eglise_id','type','activite','jour','date_prev','heure','nom_resp','nom_pasteur','date_valid']
        for k in required:
            if not (data.get(k) or '').strip():
                return JsonResponse({'success': False, 'error': f"Champ requis manquant: {k}"}, status=400)
        eg = get_object_or_404(Eglise, id=int(data.get('eglise_id')))
        from datetime import datetime as _dt
        dprev = _dt.strptime(data.get('date_prev'), '%Y-%m-%d').date()
        heure = _dt.strptime(data.get('heure'), '%H:%M').time()
        dvalid = _dt.strptime(data.get('date_valid'), '%Y-%m-%d').date()
        budget = None
        if str(data.get('budget') or '').strip():
            try:
                budget = Decimal(str(data.get('budget')))
            except Exception:
                budget = None
        obj = StrategicPlan.objects.create(
            eglise=eg,
            departement=(data.get('departement') or '').strip(),
            type_planification=(data.get('type') or '').strip(),
            activite=(data.get('activite') or '').strip(),
            jour=(data.get('jour') or '').strip(),
            date_prev=dprev,
            heure=heure,
            lieu=(data.get('lieu') or '').strip(),
            besoins=(data.get('besoins') or '').strip(),
            responsable=(data.get('responsable') or '').strip(),
            objectif=(data.get('objectif') or '').strip(),
            budget=budget,
            statut=(data.get('statut') or 'À venir').strip() or 'À venir',
            nom_resp=(data.get('nom_resp') or '').strip(),
            nom_pasteur=(data.get('nom_pasteur') or '').strip(),
            date_valid=dvalid,
        )
        return JsonResponse({'success': True, 'item': _serialize_sp(obj)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@require_http_methods(["POST"]) 
def api_sp_update(request, item_id: int):
    data = _parse_json(request)
    try:
        obj = StrategicPlan.objects.get(id=item_id)
        if 'eglise_id' in data:
            try:
                obj.eglise = Eglise.objects.get(id=int(data.get('eglise_id')))
            except Exception:
                pass
        if 'departement' in data: obj.departement = (data.get('departement') or '').strip()
        if 'type' in data: obj.type_planification = (data.get('type') or '').strip()
        if 'activite' in data: obj.activite = (data.get('activite') or '').strip()
        if 'jour' in data: obj.jour = (data.get('jour') or '').strip()
        if 'date_prev' in data:
            try:
                from datetime import datetime as _dt
                obj.date_prev = _dt.strptime(data.get('date_prev') or '', '%Y-%m-%d').date()
            except Exception:
                pass
        if 'heure' in data:
            try:
                from datetime import datetime as _dt
                obj.heure = _dt.strptime(data.get('heure') or '', '%H:%M').time()
            except Exception:
                pass
        if 'lieu' in data: obj.lieu = (data.get('lieu') or '').strip()
        if 'besoins' in data: obj.besoins = (data.get('besoins') or '').strip()
        if 'responsable' in data: obj.responsable = (data.get('responsable') or '').strip()
        if 'objectif' in data: obj.objectif = (data.get('objectif') or '').strip()
        if 'budget' in data:
            try:
                obj.budget = Decimal(str(data.get('budget')))
            except Exception:
                obj.budget = None
        if 'statut' in data: obj.statut = (data.get('statut') or 'À venir').strip() or 'À venir'
        if 'nom_resp' in data: obj.nom_resp = (data.get('nom_resp') or '').strip()
        if 'nom_pasteur' in data: obj.nom_pasteur = (data.get('nom_pasteur') or '').strip()
        if 'date_valid' in data:
            try:
                from datetime import datetime as _dt
                obj.date_valid = _dt.strptime(data.get('date_valid') or '', '%Y-%m-%d').date()
            except Exception:
                pass
        obj.save()
        return JsonResponse({'success': True, 'item': _serialize_sp(obj)})
    except StrategicPlan.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Enregistrement introuvable.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@require_http_methods(["POST"]) 
def api_sp_delete(request, item_id: int):
    try:
        obj = StrategicPlan.objects.get(id=item_id)
        obj.delete()
        return JsonResponse({'success': True})
    except StrategicPlan.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Enregistrement introuvable.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@require_http_methods(["GET"])
def api_sp_export_csv(request):
    try:
        qs = StrategicPlan.objects.select_related('eglise').all().order_by('-date_prev','heure','activite')
        eg = request.GET.get('eglise_id')
        if eg:
            try:
                qs = qs.filter(eglise_id=int(eg))
            except Exception:
                pass
        # Filtres de période
        df = (request.GET.get('date_from') or '').strip()
        dt = (request.GET.get('date_to') or '').strip()
        if df:
            try:
                from datetime import datetime as _dt
                qs = qs.filter(date_prev__gte=_dt.strptime(df, '%Y-%m-%d').date())
            except Exception:
                pass
        if dt:
            try:
                from datetime import datetime as _dt
                qs = qs.filter(date_prev__lte=_dt.strptime(dt, '%Y-%m-%d').date())
            except Exception:
                pass
        q = (request.GET.get('q') or '').strip()
        if q:
            from django.db.models import Q
            qs = qs.filter(
                Q(activite__icontains=q) | Q(departement__icontains=q) | Q(lieu__icontains=q)
                | Q(besoins__icontains=q) | Q(responsable__icontains=q) | Q(objectif__icontains=q)
                | Q(nom_resp__icontains=q) | Q(nom_pasteur__icontains=q)
            )
        import csv
        # Réponse CSV + BOM UTF-8 (pour Excel Windows)
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="planification_strategique.csv"'
        response.write('\ufeff')  # BOM
        w = csv.writer(response, delimiter=';')
        # Helpers accents
        def clean(v):
            return _fix_mojibake(_to_str(v))
        # En-têtes alignés avec le tableau HTML (hors colonne "Actions")
        w.writerow([
            'Église','Département','Type','Activité/Événement','Jour','Date','Heure','Lieu','Besoins','Responsable','Objectif','Budget','Statut','Validé par','Date validation'
        ])
        for r in qs:
            eglise_nom = clean(getattr(getattr(r,'eglise',None),'nom',''))
            departement = clean(r.departement)
            type_plan = clean(getattr(r, 'type_planification', ''))
            activite = clean(r.activite)
            jour = clean(r.jour)
            date_prev = r.date_prev.strftime('%Y-%m-%d') if r.date_prev else ''
            heure = r.heure.strftime('%H:%M') if r.heure else ''
            lieu = clean(r.lieu)
            besoins = clean(r.besoins)
            responsable = clean(r.responsable)
            objectif = clean(r.objectif)
            budget = str(r.budget) if r.budget is not None else ''
            statut = clean(r.statut)
            nom_resp = clean(r.nom_resp)
            nom_pasteur = clean(r.nom_pasteur)
            date_valid = r.date_valid.strftime('%Y-%m-%d') if r.date_valid else ''
            valide_par = (nom_resp + ' / ' + nom_pasteur).strip(' / ')
            w.writerow([
                eglise_nom, departement, type_plan, activite, jour,
                date_prev, heure, lieu, besoins, responsable, objectif, budget, statut,
                valide_par, date_valid
            ])
        return response
    except Exception as e:
        return HttpResponse(f"Erreur export CSV: {e}", status=500)

@require_http_methods(["GET"])
def api_sp_export_xlsx(request):
    try:
        qs = StrategicPlan.objects.select_related('eglise').all().order_by('-date_prev','heure','activite')
        eg = request.GET.get('eglise_id')
        if eg:
            try:
                qs = qs.filter(eglise_id=int(eg))
            except Exception:
                pass
        # Filtres de période
        df = (request.GET.get('date_from') or '').strip()
        dt = (request.GET.get('date_to') or '').strip()
        if df:
            try:
                from datetime import datetime as _dt
                qs = qs.filter(date_prev__gte=_dt.strptime(df, '%Y-%m-%d').date())
            except Exception:
                pass
        if dt:
            try:
                from datetime import datetime as _dt
                qs = qs.filter(date_prev__lte=_dt.strptime(dt, '%Y-%m-%d').date())
            except Exception:
                pass
        q = (request.GET.get('q') or '').strip()
        if q:
            from django.db.models import Q
            qs = qs.filter(
                Q(activite__icontains=q) | Q(departement__icontains=q) | Q(lieu__icontains=q)
                | Q(besoins__icontains=q) | Q(responsable__icontains=q) | Q(objectif__icontains=q)
                | Q(nom_resp__icontains=q) | Q(nom_pasteur__icontains=q)
            )
        try:
            from openpyxl import Workbook
        except Exception:
            # Fallback vers CSV Excel-compatible
            return api_sp_export_csv(request)
        wb = Workbook(); ws = wb.active; ws.title = 'Planif'
        headers = ['Église','Département','Type','Activité/Événement','Jour','Date','Heure','Lieu','Besoins','Responsable','Objectif','Budget','Statut','Nom Responsable','Nom Pasteur','Date Validation']
        ws.append(headers)
        for r in qs:
            ws.append([
                getattr(getattr(r,'eglise',None),'nom',''), r.departement, r.type_planification, r.activite, r.jour,
                r.date_prev.strftime('%Y-%m-%d') if r.date_prev else '', r.heure.strftime('%H:%M') if r.heure else '', r.lieu,
                r.besoins, r.responsable, r.objectif, (str(r.budget) if r.budget is not None else ''), r.statut,
                r.nom_resp, r.nom_pasteur, r.date_valid.strftime('%Y-%m-%d') if r.date_valid else ''
            ])
        for col in ws.columns:
            max_len=0; letter=col[0].column_letter
            for c in col:
                v=str(c.value) if c.value is not None else ''
                if len(v)>max_len: max_len=len(v)
            ws.column_dimensions[letter].width=min(40,max(10,max_len+2))
        from io import BytesIO
        bio=BytesIO(); wb.save(bio); bio.seek(0)
        resp=HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition']='attachment; filename="planification_strategique.xlsx"'
        return resp
    except Exception as e:
        return HttpResponse(f"Erreur export XLSX: {e}", status=500)

def test_image(request):
    """Vue pour tester l'affichage d'une image"""
    return render(request, 'main/test_image.html')

@require_POST
def supprimer_toutes_finances(request):
    """Supprime définitivement toutes les opérations financières (table Finance). Retour JSON."""
    try:
        count = Finance.objects.count()
        Finance.objects.all().delete()
        return JsonResponse({ 'success': True, 'deleted_count': count })
    except Exception as e:
        return JsonResponse({ 'success': False, 'error': str(e) }, status=400)

@login_required
@ensure_csrf_cookie
def finances_records_all(request):
    """Affiche un tableau dédié avec tous les enregistrements financiers."""
    operations = (
        Finance.objects
        .select_related('eglise', 'report')
        .prefetch_related('report__attachments')
        .all()
        .order_by('-date_operation', '-id')
    )
    return render(request, 'main/finances_records_all.html', {
        'operations': operations,
    })


def finance_report_detail(request, report_id: int):
    """Affiche le détail d'un FinanceReport: en-tête, lignes associées et pièces jointes."""
    report = get_object_or_404(FinanceReport.objects.select_related('eglise'), id=report_id)
    
    # 1) Lignes liées directement par report_id (via related_name='operations')
    lignes = list(report.operations.select_related('eglise').order_by('date_operation', 'id'))
    
    # 2) Fallbacks pour anciennes données (sans report)
    if not lignes:
        qs = Finance.objects.select_related('eglise').filter(eglise=report.eglise)
        # 2a) Filtrer par période si disponible
        if report.periode_du and report.periode_au:
            qs = qs.filter(date_operation__gte=report.periode_du, date_operation__lte=report.periode_au)
        # 2b) Sinon, tenter par date_rapport exacte
        elif report.date_rapport:
            qs = qs.filter(date_operation=report.date_rapport)
        # 2c) Sinon, récupérer toutes les opérations de l'église sans rapport assigné
        else:
            qs = qs.filter(report__isnull=True)
        lignes = list(qs.order_by('date_operation', 'id'))

    # Pièces jointes du rapport
    try:
        attachments = list(FinanceAttachment.objects.filter(report_id=report.id))
    except Exception:
        attachments = []

    return render(request, 'main/finance_report_detail.html', {
        'report': report,
        'lignes': lignes,
        'attachments': attachments,
    })

def supprimer_eglise(request, eglise_id):
    """Vue pour supprimer une église"""
    if request.method == 'POST':
        try:
            eglise = get_object_or_404(Eglise, id=eglise_id)
            nom_eglise = eglise.nom
            eglise.delete()
            messages.success(request, f"L'église '{nom_eglise}' a été supprimée avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de la suppression : {str(e)}")
    
    return redirect('main:eglises')
